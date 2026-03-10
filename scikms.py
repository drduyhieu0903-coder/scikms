"""
SciKMS — Scientific Knowledge Management System
================================================
Run: streamlit run app.py
Requirements: pip install -r requirements.txt
"""

import streamlit as st
import sqlite3
import os
import re
import json
import shutil
import hashlib
import unicodedata
import zipfile
import io
from datetime import datetime
from pathlib import Path
import pandas as pd
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# ── Optional dependencies ───────────────────────────────────────
try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ══════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG & CUSTOM CSS
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="SciKMS · Medical Literature Management",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded",
)

CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=Lora:ital,wght@0,400;0,600;0,700;1,400&family=Inter:wght@300;400;500;600;700&display=swap');

/* ─── Global Background (Light Mode) ────────────────────────── */
html, body, [data-testid="stAppViewContainer"] {
    background: #f8fafc !important;
    color: #1e293b !important;
}

/* ─── Optimize page width (Works smoothly with wide layout) ─── */
[data-testid="block-container"] {
    max-width: 100% !important;
    padding-left: 3rem !important;
    padding-right: 3rem !important;
    padding-top: 2rem !important;
}

/* ─── Customize Sidebar ─── */
[data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 1px solid #e2e8f0 !important;
}
/* Deleted forced width part [aria-expanded="true"] để Streamlit tự co giãn */



/* ─── Prevent Material Icon layout break ─── */
span.material-icons, span.material-symbols-rounded {
    font-family: 'Material Icons', 'Material Symbols Rounded', sans-serif !important;
    font-size: 20px !important;
    line-height: 1 !important;
    display: inline-block !important;
    vertical-align: middle !important;
}

/* ─── Typography ───────────────────────────────────────── */
h1, h2, h3 { font-family: 'Lora', serif !important; color: #334155 !important; }
/* Deleted div and span để không làm hỏng icon của Streamlit */
p, label, li, .stMarkdown { font-family: 'Inter', sans-serif !important; }
code, .mono { font-family: 'IBM Plex Mono', monospace !important; }

/* ─── Buttons ──────────────────────────────────────────── */
.stButton > button {
    background: linear-gradient(135deg, #c9a84c, #e6bf70) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-family: 'Inter', sans-serif !important;
    padding: 0.5rem 1.2rem !important;
    transition: all 0.2s !important;
    box-shadow: 0 2px 4px rgba(201,168,76,0.15) !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 15px rgba(201,168,76,0.25) !important;
}
.stButton > button[kind="secondary"] {
    background: #ffffff !important;
    color: #64748b !important;
    border: 1px solid #e2e8f0 !important;
}
.stButton > button[kind="secondary"]:hover {
    color: #1e293b !important;
    border-color: #c9a84c !important;
    background: #fdfdfd !important;
}

/* ─── Input / Select ────────────────────────────────────── */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stSelectbox > div > div {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    color: #1e293b !important;
    border-radius: 8px !important;
    font-family: 'Inter', sans-serif !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: #c9a84c !important;
    box-shadow: 0 0 0 2px rgba(201,168,76,0.1) !important;
}

/* ─── Cards & Expanders (Light background, soft shadow) ────────── */
.paper-card {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 14px;
    padding: 24px;
    margin-bottom: 20px;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    box-shadow: 0 1px 3px rgba(0,0,0,0.02);
}
.paper-card:hover {
    border-color: #c9a84c44;
    box-shadow: 0 10px 30px rgba(0,0,0,0.06);
    transform: translateY(-2px);
}
.paper-card.starred { border-left: 5px solid #c9a84c; }

.paper-title {
    font-family: 'Lora', serif;
    font-size: 16px;
    font-weight: 700;
    color: #1e293b;
    line-height: 1.4;
    margin-bottom: 8px;
}
.gold-text { color: #c9a84c; font-weight: bold; }
.doi-text { color: #15803d; font-family: 'IBM Plex Mono', monospace; font-size: 12px; }

/* ─── Tags & Badges ─────────────────────────────────────── */
.tag, .tag-badge {
    display: inline-block;
    background: #f1f5f9;
    color: #475569;
    border: 1px solid #e2e8f0;
    border-radius: 6px;
    padding: 3px 10px;
    font-size: 11px;
    font-family: 'Inter', sans-serif;
    margin: 3px;
    font-weight: 500;
    white-space: nowrap;
}
.tag.highlight { background: #fef9c3; color: #854d0e; border-color: #fde047; }
.badge-read    { color: #15803d; font-size: 11px; font-family: 'IBM Plex Mono', monospace; font-weight: 600; }
.badge-reading { color: #b45309; font-size: 11px; font-family: 'IBM Plex Mono', monospace; font-weight: 600; }
.badge-unread  { color: #94a3b8; font-size: 11px; font-family: 'IBM Plex Mono', monospace; }

/* ─── Metric cards ──────────────────────────────────────── */
.metric-box {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 18px;
    text-align: center;
    box-shadow: 0 1px 2px rgba(0,0,0,0.02);
}
.metric-val {
    font-size: 26px;
    font-weight: 700;
    font-family: 'IBM Plex Mono', monospace;
    color: #c9a84c;
}
.metric-label {
    font-size: 11px;
    color: #94a3b8;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-top: 6px;
}

/* ─── Abstract preview (Nền xanh nhạt thanh lịch) ────────── */
.abstract-container {
    margin: 15px 0;
    position: relative;
}

.abstract-box {
    background: #f8fafc;
    border-left: 4px solid #c9a84c;
    border-radius: 4px;
    padding: 16px 20px;
    font-size: 14px;
    line-height: 1.8;
    color: #475569;
    font-family: 'Lora', serif;
    /* CSS Line Clamp mặc định 2 dòng */
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
    overflow: hidden;
    text-overflow: ellipsis;
    transition: all 0.3s ease;
}

.abstract-toggle {
    display: none;
}

.abstract-toggle:checked + .abstract-box {
    -webkit-line-clamp: unset;
}

.abstract-label {
    display: block;
    text-align: center;
    font-size: 12px;
    color: #c9a84c;
    cursor: pointer;
    font-weight: 600;
    margin-top: 4px;
    user-select: none;
}

.abstract-label:hover {
    text-decoration: underline;
}

.abstract-toggle:checked ~ .abstract-label::after {
    content: "LESS ▲";
}

.abstract-toggle:not(:checked) ~ .abstract-label::after {
    content: "MORE ▼";
}

/* ─── Rename preview ────────────────────────────────────── */
.rename-old { color: #be123c; text-decoration: line-through; font-family: 'IBM Plex Mono', monospace; font-size: 12px; opacity: 0.6; }
.rename-new { color: #15803d; font-family: 'IBM Plex Mono', monospace; font-size: 12px; font-weight: 600; }

/* ─── Search highlight ──────────────────────────────────── */
.search-match { background: #fef08a; border-radius: 3px; padding: 0 4px; color: #854d0e; font-weight: 600; }

/* ─── Divider ───────────────────────────────────────────── */
.custom-divider { border: none; border-top: 1px solid #e2e8f0; margin: 24px 0; }

/* ─── File uploader ─────────────────────────────────────── */
[data-testid="stFileUploader"] {
    background: #ffffff !important;
    border: 2px dashed #cbd5e1 !important;
    border-radius: 16px !important;
    padding: 20px !important;
}
[data-testid="stFileUploader"]:hover { border-color: #c9a84c !important; background: #fdfdfd !important; }

/* ─── Dataframe / Table ─────────────────────────────────── */
.stDataFrame { background: #ffffff !important; border-radius: 12px !important; border: 1px solid #e2e8f0 !important; }
thead tr th { background: #f8fafc !important; color: #475569 !important; font-family: 'Inter', sans-serif !important; font-size: 12px !important; font-weight: 600 !important; }
tbody tr:hover td { background: #f1f5f9 !important; }

/* ─── Expander ──────────────────────────────────────────── */
.streamlit-expanderHeader {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 10px !important;
    color: #1e293b !important;
    font-family: 'Inter', sans-serif !important;
}
.streamlit-expanderContent {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-top: none !important;
}


/* ─── Sidebar radio / selectbox ─────────────────────────── */
.stRadio > label { color: #64748b !important; font-size: 14px !important; font-weight: 500 !important; }

/* ─── Sidebar navigation buttons (larger & easier to click) ─── */
[data-testid="stSidebar"] .stButton > button {
    padding: 0.7rem 1rem !important;
    font-size: 15px !important;
    font-weight: 500 !important;
    border-radius: 10px !important;
    text-align: left !important;
    justify-content: flex-start !important;
    letter-spacing: 0.01em !important;
    transition: all 0.15s ease !important;
    min-height: 48px !important;
}
[data-testid="stSidebar"] .stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #c9a84c, #dbb85e) !important;
    color: #ffffff !important;
    border: none !important;
    box-shadow: 0 3px 10px rgba(201,168,76,0.2) !important;
}
[data-testid="stSidebar"] .stButton > button[kind="secondary"] {
    background: #f8fafc !important;
    color: #334155 !important;
    border: 1px solid #e2e8f0 !important;
}
[data-testid="stSidebar"] .stButton > button[kind="secondary"]:hover {
    background: #f1f5f9 !important;
    color: #1e293b !important;
    border-color: #c9a84c !important;
    transform: translateX(3px) !important;
}

/* ─── Progress bar ──────────────────────────────────────── */
.stProgress > div > div { background: linear-gradient(90deg,#c9a84c,#e6bf70) !important; }

/* ─── Tooltip / info ────────────────────────────────────── */
.info-tip {
    background: #eff6ff;
    border: 1px solid #bfdbfe;
    border-radius: 10px;
    padding: 12px 18px;
    font-size: 13px;
    color: #1e40af;
    line-height: 1.6;
}

/* ─── Dark Mode Overrides for UI components ──────────────── */
[data-testid="stMetricValue"] { color: #1e293b !important; }
.stMarkdown h2, .stMarkdown h3 { border-bottom: none !important; }

/* ─── Force light mode on ALL Streamlit widgets ──────────── */
/* Number input */
[data-testid="stNumberInput"] input,
[data-testid="stNumberInput"] button {
    background: #ffffff !important;
    color: #1e293b !important;
    border-color: #e2e8f0 !important;
}

/* Expander (modern Streamlit versions) */
[data-testid="stExpander"] {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 10px !important;
}
[data-testid="stExpander"] summary,
[data-testid="stExpander"] details > summary {
    background: #ffffff !important;
    color: #1e293b !important;
}
[data-testid="stExpander"] [data-testid="stExpanderDetails"] {
    background: #ffffff !important;
}

/* File uploader inner text & button */
[data-testid="stFileUploader"] label,
[data-testid="stFileUploader"] span,
[data-testid="stFileUploader"] small,
[data-testid="stFileUploader"] section {
    color: #475569 !important;
}
[data-testid="stFileUploader"] button {
    background: #f1f5f9 !important;
    color: #1e293b !important;
    border: 1px solid #e2e8f0 !important;
}
[data-testid="stFileUploaderDropzone"] {
    background: #ffffff !important;
    color: #475569 !important;
}
[data-testid="stFileUploaderDropzoneInstructions"] span,
[data-testid="stFileUploaderDropzoneInstructions"] div,
[data-testid="stFileUploaderDropzoneInstructions"] small {
    color: #64748b !important;
}

/* Selectbox dropdown menu */
[data-testid="stSelectbox"] div[role="listbox"] {
    background: #ffffff !important;
    color: #1e293b !important;
}

/* Checkbox & Radio */
.stCheckbox label span,
.stRadio div label { color: #1e293b !important; }

/* General: all text-like elements */
[data-testid="stAppViewContainer"] label { color: #475569 !important; }
[data-testid="stAppViewContainer"] .stAlert { border-radius: 10px !important; }

/* Sidebar labels */
[data-testid="stSidebar"] label { color: #475569 !important; }
[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] { background: #f8fafc !important; }
[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] * { color: #1e293b !important; }

/* ─── Hide Streamlit branding ─── */
#MainMenu { visibility: hidden !important; }
footer { visibility: hidden !important; }
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)





# ══════════════════════════════════════════════════════════════════════════════
#  DATABASE
# ══════════════════════════════════════════════════════════════════════════════
DB_PATH = "scikms.db"
STORAGE_DIR = Path("storage")
STORAGE_DIR.mkdir(exist_ok=True)


def get_db():
    """Trả về SQLite connection. Hỗ trợ context manager (with get_db() as conn)."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS papers (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            md5         TEXT UNIQUE,
            original_filename TEXT,
            renamed_filename  TEXT,
            title       TEXT,
            authors     TEXT,
            year        INTEGER,
            journal     TEXT,
            doi         TEXT,
            abstract    TEXT,
            keywords    TEXT,
            full_text   TEXT,
            tags        TEXT,
            notes       TEXT DEFAULT '',
            highlights  TEXT DEFAULT '[]',
            status      TEXT DEFAULT 'unread',
            starred     INTEGER DEFAULT 0,
            pages       INTEGER DEFAULT 0,
            added_at    TEXT,
            file_path   TEXT,
            project     TEXT DEFAULT ''
        );

        CREATE VIRTUAL TABLE IF NOT EXISTS papers_fts
            USING fts5(
                id UNINDEXED,
                title,
                authors,
                abstract,
                keywords,
                full_text,
                content='papers',
                content_rowid='id'
            );

        CREATE TRIGGER IF NOT EXISTS papers_ai AFTER INSERT ON papers BEGIN
            INSERT INTO papers_fts(rowid,id,title,authors,abstract,keywords,full_text)
            VALUES (new.id,new.id,new.title,new.authors,new.abstract,new.keywords,new.full_text);
        END;
        CREATE TRIGGER IF NOT EXISTS papers_ad AFTER DELETE ON papers BEGIN
            INSERT INTO papers_fts(papers_fts,rowid,id,title,authors,abstract,keywords,full_text)
            VALUES ('delete',old.id,old.id,old.title,old.authors,old.abstract,old.keywords,old.full_text);
        END;
        CREATE TRIGGER IF NOT EXISTS papers_au AFTER UPDATE ON papers BEGIN
            INSERT INTO papers_fts(papers_fts,rowid,id,title,authors,abstract,keywords,full_text)
            VALUES ('delete',old.id,old.id,old.title,old.authors,old.abstract,old.keywords,old.full_text);
            INSERT INTO papers_fts(rowid,id,title,authors,abstract,keywords,full_text)
            VALUES (new.id,new.id,new.title,new.authors,new.abstract,new.keywords,new.full_text);
        END;
    """)
    conn.commit()

    # Migrate table
    try:
        c.execute("ALTER TABLE papers ADD COLUMN reading_position INTEGER DEFAULT 0")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    try:
        c.execute("ALTER TABLE papers ADD COLUMN project TEXT DEFAULT ''")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    conn.close()


def get_config_path() -> Path:
    return STORAGE_DIR / "scikms_config.json"

def read_config() -> dict:
    conf = get_config_path()
    if conf.exists():
        try:
            with open(conf, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError, ValueError):
            pass  # File hỏng hoặc không đọc được — trả về default
    return {"has_seen_tutorial": False}

def save_config(config: dict):
    with open(get_config_path(), "w", encoding="utf-8") as f:
        json.dump(config, f)

# ─── Gamified Onboarding Quest ───────────────────────────────────────────────
def get_quest_state() -> int:
    # 0 = not started, 1-7 = ongoing, 99 = finished
    if read_config().get("has_seen_tutorial", False):
        return 99
    if "quest_step" not in st.session_state:
        st.session_state.quest_step = 1
    return st.session_state.quest_step

def finish_quest():
    conf = read_config()
    conf["has_seen_tutorial"] = True
    save_config(conf)
    st.session_state.quest_step = 99
    st.rerun()

def render_quest_tracker():
    step = get_quest_state()
    if step == 99:
        return
        
    quest_data = {
        1: {
            "title": "Module 1: Getting Started",
            "desc": "Welcome to SciKMS! This interactive guide will teach you the basics. To begin, please click the **📚 Library** tab located at the top navigation bar.",
            "check": lambda: st.session_state.active_tab == "📚 Library"
        },
        2: {
            "title": "Module 2: The Import System",
            "desc": "Great! The Library is where all your papers live. Now, let's learn how to add a new paper. Click the **⬆️ Import** tab.",
            "check": lambda: st.session_state.active_tab == "⬆️ Import"
        },
        3: {
            "title": "Module 3: Uploading a Paper",
            "desc": "You are now in the Import Center. Drag and drop a PDF file into the dashed box (or click to browse). Then, click the blue **🚀 Process PDF files** button. The AI will automatically extract the Title, Authors, Abstract, and DOI for you.",
            "check": lambda: len(get_all_papers()) > 0 and st.session_state.active_tab == "📚 Library"
        },
        4: {
            "title": "Module 4: Searching the Knowledge Base",
            "desc": "Excellent! Your paper has been added. Now, let's try finding it. Click the **🔍 Search** tab.",
            "check": lambda: st.session_state.active_tab == "🔍 Search"
        },
        5: {
            "title": "Module 5: Organizing and Renaming",
            "desc": "You can search by keyword, author, or even the full text here! Next, let's see how to clean up messy file names. Click the **✏️ Rename ** tab.",
            "check": lambda: st.session_state.active_tab == "✏️ Rename "
        },
        6: {
            "title": "Module 6: Exporting Citations",
            "desc": "Here, you can click 'Apply auto-rename' to automatically rename all your physical PDF files to a standard format (e.g., '[2023] Author - Title.pdf'). Now, click the **📤 Export** tab.",
            "check": lambda: st.session_state.active_tab == "📤 Export"
        },
        7: {
            "title": "Module 7: Reading and Annotating",
            "desc": "In the Export tab, you can download your library as an Excel sheet or as RIS/BibTeX files for Zotero/EndNote. Finally, let's read a paper. Go back to the **📚 Library** tab, and click the **📖 Quick read** button on any paper card to open the built-in PDF reader.",
            "check": lambda: st.session_state.active_tab == "📚 Library" and st.session_state.selected_paper_id is not None
        }
    }
    
    # Auto-advance logic
    if step in quest_data and quest_data[step]["check"]():
        st.session_state.quest_step += 1
        st.rerun()
        return

    # Special case: Finish on step 8
    if step > 7:
        st.toast("🎉 Tutorial Quest Completed! Welcome to SciKMS.")
        finish_quest()
        return

    # Render banner
    current_mission = quest_data[step]
    
    st.markdown(f"""
    <div style="background: #eef2ff; border-left: 5px solid #6366f1; padding: 15px 20px; border-radius: 6px; margin-bottom: 25px; display: flex; justify-content: space-between; align-items: center;">
        <div>
            <div style="font-size: 13px; font-weight: 700; color: #4338ca; margin-bottom: 4px; text-transform: uppercase; letter-spacing: 0.5px;">🎓 Beginner Tutorial ({step}/7)</div>
            <div style="font-size: 16px; color: #1e293b; font-weight: 600;">{current_mission['title']}</div>
            <div style="font-size: 14px; color: #475569; margin-top: 5px;">👉 {current_mission['desc']}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1, 6])
    with col1:
        if st.button("⏭️ Skip Tutorial", key="skip_quest"):
            finish_quest()
    with col2:
        if st.button("🔄 Reset Quest", key="reset_quest"):
            st.session_state.quest_step = 1
            st.rerun()

init_db()

# ══════════════════════════════════════════════════════════════════════════════
#  FILE PROCESSING UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def md5_of_bytes(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()


def slugify(text: str, max_len: int = 60) -> str:
    """Normalize string to safe filename."""
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s\-]", "", text)
    text = re.sub(r"[\s_]+", "_", text).strip("_")
    return text[:max_len]


def build_renamed_filename(paper: dict) -> str:
    """
    Format: [YYYY] LastName - Short_Title.pdf
    Example: [2023] Zhang_Nguyen - 3D_Photogrammetric_Rhinoplasty.pdf
    """
    year = paper.get("year") or "XXXX"
    raw_authors = paper.get("authors") or ""
    title = paper.get("title") or "Untitled"

    # Extract author last names
    author_parts = [a.strip() for a in raw_authors.split(";") if a.strip()]
    if author_parts:
        last_names = []
        for a in author_parts[:2]:
            # Format "Last, First" or "First Last"
            parts = a.split(",")
            last_names.append(slugify(parts[0].strip(), 20))
        author_str = "_".join(last_names)
    else:
        author_str = "Unknown"

    # Rút 5-6 from đầu tiêu đề
    title_words = re.sub(r'[^\w\s]', '', title).split()
    short_title = "_".join(title_words[:6])
    short_title = slugify(short_title, 50)

    return f"[{year}] {author_str} - {short_title}.pdf"


def rename_physical_file(old_path: str, new_filename: str) -> str:
    """
    Rename files vật lý trên ổ cứng and trả về đường dẫn mới.
    Includes collision handling.
    """
    if not old_path or not os.path.exists(old_path):
        return old_path  # Trả về đường dẫn cũ nếu file không tồn tại

    # Get directory containing current file (usually "storage")
    directory = os.path.dirname(old_path)
    
    # Create new path
    new_path = os.path.join(directory, new_filename)

    # Prevent filename collisions (if 2 papers have same year, author, title)
    if os.path.exists(new_path) and new_path != old_path:
        base_name, ext = os.path.splitext(new_filename)
        counter = 1
        while os.path.exists(os.path.join(directory, f"{base_name}_{counter}{ext}")):
            counter += 1
        new_path = os.path.join(directory, f"{base_name}_{counter}{ext}")

    # Rename files thực tế
    if new_path != old_path:
        try:
            os.rename(old_path, new_path)
        except Exception as e:
            st.error(f"Error renaming file on disk: {e}")
            return old_path

    return new_path


def extract_doi(text: str) -> str | None:
    pattern = r'\b10\.\d{4,}/[^\s"\'<>]{3,}\b'
    m = re.search(pattern, text)
    return m.group(0).rstrip(".,;") if m else None


def fetch_crossref(doi: str) -> dict:
    if not HAS_REQUESTS:
        return {}
    try:
        url = f"https://api.crossref.org/works/{doi}"
        r = requests.get(url, timeout=8,
                         headers={"User-Agent": "SciKMS/1.0 (mailto:user@example.com)"})
        if r.status_code == 200:
            msg = r.json().get("message", {})
            authors = "; ".join(
                f"{a.get('family','')}, {a.get('given','')}".strip(", ")
                for a in msg.get("author", [])
            )
            issued = msg.get("issued", {}).get("date-parts", [[None]])[0]
            year = issued[0] if issued else None
            return {
                "title": (msg.get("title") or [""])[0],
                "authors": authors,
                "year": year,
                "journal": (msg.get("container-title") or [""])[0],
                "doi": doi,
                "abstract": re.sub(r'<[^>]+>', '', msg.get("abstract") or ""),
            }
    except Exception:
        pass
    return {}


def fetch_pubmed(query: str = "", pmid: str = "", doi: str = "") -> dict:
    """
    Search papers qua PubMed E-utilities API (miễn phí, không cần key).
    Supports: search by PMID, DOI, or free search string.
    Returns dict like fetch_crossref().
    """
    if not HAS_REQUESTS:
        return {}
    BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
    headers = {"User-Agent": "SciKMS/1.0 (mailto:user@example.com)"}
    try:
        # ── Step 1: Get PMID ──────────────────────────────────────────────
        target_pmid = pmid
        if not target_pmid:
            # Search by DOI or query string
            search_term = f"{doi}[DOI]" if doi else query
            if not search_term:
                return {}
            r = requests.get(
                f"{BASE}/esearch.fcgi",
                params={"db": "pubmed", "term": search_term,
                        "retmax": 1, "retmode": "json"},
                timeout=8, headers=headers,
            )
            if r.status_code != 200:
                return {}
            ids = r.json().get("esearchresult", {}).get("idlist", [])
            if not ids:
                return {}
            target_pmid = ids[0]

        # ── Bước 2: Fetch chi tiết papers qua efetch ─────────────────────
        r2 = requests.get(
            f"{BASE}/efetch.fcgi",
            params={"db": "pubmed", "id": target_pmid,
                    "retmode": "xml", "rettype": "abstract"},
            timeout=10, headers=headers,
        )
        if r2.status_code != 200:
            return {}

        xml = r2.text

        def _xml_tag(tag: str, text: str) -> str:
            """Extract simple XML tag content."""
            m = re.search(fr"<{tag}[^>]*>(.*?)</{tag}>", text, re.S)
            return m.group(1).strip() if m else ""

        def _xml_tags(tag: str, text: str) -> list[str]:
            return re.findall(fr"<{tag}[^>]*>(.*?)</{tag}>", text, re.S)

        # Title
        title = re.sub(r"<[^>]+>", "", _xml_tag("ArticleTitle", xml))

        # Authors: LastName + ForeName
        author_nodes = _xml_tags("Author", xml)
        authors_list = []
        for node in author_nodes:
            last  = re.sub(r"<[^>]+>", "", _xml_tag("LastName", node))
            fore  = re.sub(r"<[^>]+>", "", _xml_tag("ForeName", node))
            initials = re.sub(r"<[^>]+>", "", _xml_tag("Initials", node))
            if last:
                authors_list.append(f"{last}, {fore or initials}".strip(", "))
        authors = "; ".join(authors_list)

        # Year
        year_str = _xml_tag("PubDate", xml)
        year_m   = re.search(r"\b(19|20)\d{2}\b", year_str)
        year     = int(year_m.group(0)) if year_m else None

        # Journal
        journal = re.sub(r"<[^>]+>", "", _xml_tag("Title", xml))  # Journal Title tag
        if not journal:
            journal = re.sub(r"<[^>]+>", "", _xml_tag("ISOAbbreviation", xml))

        # Abstract
        abstract_parts = _xml_tags("AbstractText", xml)
        abstract = " ".join(re.sub(r"<[^>]+>", "", p).strip() for p in abstract_parts)

        # DOI from ArticleId
        doi_found = ""
        for aid in _xml_tags("ArticleId", xml):
            if 'IdType="doi"' in xml:
                doi_found_m = re.search(r'<ArticleId IdType="doi">(.*?)</ArticleId>', xml)
                if doi_found_m:
                    doi_found = doi_found_m.group(1).strip()
                    break

        # Keywords MeSH
        mesh_terms = _xml_tags("DescriptorName", xml)
        keywords = ", ".join(re.sub(r"<[^>]+>", "", t) for t in mesh_terms[:10])

        if not title:
            return {}

        return {
            "title":    title,
            "authors":  authors,
            "year":     year,
            "journal":  journal,
            "doi":      doi_found or doi,
            "abstract": abstract,
            "keywords": keywords,
            "pmid":     target_pmid,
        }
    except Exception:
        pass
    return {}


def import_by_pmid(pmid: str, auto_download_pdf: bool = False) -> dict:
    """Import papers from PMID ando thư viện."""
    pmid = pmid.strip()
    if not pmid.isdigit():
        return {"error": "PMID must be an integer (e.g., 17532924)"}

    # Check duplicates
    conn = get_db()
    try:
        # PubMed does not store pmid directly, check via notes or doi
        exists = conn.execute(
            "SELECT id,title FROM papers WHERE notes LIKE ?", (f"%PMID:{pmid}%",)
        ).fetchone()
    finally:
        conn.close()
    if exists:
        return {"error": f"PMID already exists: «{exists['title']}»"}

    meta = fetch_pubmed(pmid=pmid)
    if not meta or not meta.get("title"):
        return {"error": f"Metadata not found for PMID: {pmid}"}

    # If DOI exists, try using Crossref to get more
    doi = meta.get("doi", "")
    if doi:
        crossref_meta = fetch_crossref(doi)
        if crossref_meta.get("abstract") and not meta.get("abstract"):
            meta["abstract"] = crossref_meta["abstract"]

    file_path = ""
    pages = 0
    full_text = f"{meta.get('title','')} {meta.get('abstract','')}"
    pdf_url_found = None
    pdf_source = ""

    if auto_download_pdf:
        oa = find_open_access_pdf(doi=doi, title=meta.get("title",""))
        if not oa["found"]:
            # Try adding via PMC
            oa_pmc = fetch_pmc_pdf(doi=doi, pmid=pmid)
            if oa_pmc:
                oa = {"url": oa_pmc, "source": "PubMed Central", "found": True}
        if oa["found"]:
            pdf_url_found = oa["url"]
            pdf_source    = oa["source"]
            dl = download_and_save_pdf(pdf_url_found, doi=doi,
                                       filename_hint=re.sub(r'[^\w]', '_', meta.get("title",""))[:30])
            if dl["success"]:
                file_path = dl["file_path"]
                pages     = dl.get("pages", 0)
                if dl.get("full_text"):
                    full_text = dl["full_text"]

    if file_path and os.path.exists(file_path):
        with open(file_path, "rb") as _f:
            md5 = hashlib.md5(_f.read()).hexdigest()
    else:
        md5 = hashlib.md5((doi or pmid).encode()).hexdigest()

    tags = auto_tag(full_text[:5000], meta.get("keywords",""), meta.get("abstract",""))

    paper = {
        "md5": md5,
        "original_filename": f"(PMID import: {pmid})",
        "title": meta.get("title",""),
        "authors": meta.get("authors",""),
        "year": meta.get("year") or datetime.now().year,
        "journal": meta.get("journal",""),
        "doi": doi,
        "abstract": meta.get("abstract",""),
        "keywords": meta.get("keywords",""),
        "full_text": full_text,
        "tags": json.dumps(tags),
        "status": "unread",
        "starred": 0,
        "pages": pages,
        "added_at": datetime.now().strftime("%Y-%m-%d"),
        "file_path": file_path,
        "notes": f"PMID:{pmid}",
        "highlights": "[]",
        "project": "",
    }
    paper["renamed_filename"] = build_renamed_filename(paper)

    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO papers
            (md5,original_filename,renamed_filename,title,authors,year,journal,doi,
             abstract,keywords,full_text,tags,notes,highlights,status,starred,pages,added_at,file_path,project)
            VALUES (:md5,:original_filename,:renamed_filename,:title,:authors,:year,:journal,:doi,
                    :abstract,:keywords,:full_text,:tags,:notes,:highlights,:status,:starred,:pages,:added_at,:file_path,:project)
        """, paper)
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return {"error": "Paper already exists (duplicate MD5)."}
    finally:
        conn.close()

    paper["_pdf_url"]        = pdf_url_found or ""
    paper["_pdf_downloaded"] = bool(file_path)
    paper["_pdf_source"]     = pdf_source
    return paper


def fetch_unpaywall(doi: str, email: str = "scikms@example.com") -> str | None:
    """Find free Open Access PDF link via Unpaywall API. Returns PDF URL if available."""
    if not HAS_REQUESTS or not doi:
        return None
    try:
        url = f"https://api.unpaywall.org/v2/{doi}?email={email}"
        r = requests.get(url, timeout=10,
                         headers={"User-Agent": "SciKMS/1.0"})
        if r.status_code == 200:
            data = r.json()
            best = data.get("best_oa_location") or {}
            pdf_url = best.get("url_for_pdf") or best.get("url")
            if pdf_url:
                return pdf_url
            for loc in data.get("oa_locations", []):
                if loc.get("url_for_pdf"):
                    return loc["url_for_pdf"]
    except Exception:
        pass
    return None


def fetch_semantic_scholar_pdf(doi: str = "", title: str = "") -> str | None:
    """
    Find Open Access PDF link via Semantic Scholar API.
    Hỗ trợ tìm theo DOI hoặc tiêu đề papers.
    """
    if not HAS_REQUESTS:
        return None
    try:
        headers = {"User-Agent": "SciKMS/1.0"}
        # Prioritize searching by DOI
        if doi:
            url = f"https://api.semanticscholar.org/graph/v1/paper/DOI:{doi}?fields=openAccessPdf,externalIds"
            r = requests.get(url, timeout=10, headers=headers)
            if r.status_code == 200:
                data = r.json()
                oa = data.get("openAccessPdf") or {}
                if oa.get("url"):
                    return oa["url"]
        # Fallback: search by title
        if title:
            search_url = "https://api.semanticscholar.org/graph/v1/paper/search"
            params = {"query": title[:100], "fields": "openAccessPdf,title", "limit": 1}
            r = requests.get(search_url, params=params, timeout=10, headers=headers)
            if r.status_code == 200:
                items = r.json().get("data", [])
                if items:
                    oa = (items[0].get("openAccessPdf") or {})
                    if oa.get("url"):
                        return oa["url"]
    except Exception:
        pass
    return None


def fetch_pmc_pdf(doi: str = "", pmid: str = "") -> str | None:
    """
    Tìm link PDF miễn phí from PubMed Central (PMC).
    Trả về URL tải PDF nếu papers có trong PMC.
    """
    if not HAS_REQUESTS:
        return None
    try:
        headers = {"User-Agent": "SciKMS/1.0"}
        # Lấy PMCID from DOI hoặc PMID
        if doi:
            conv_url = (
                f"https://www.ncbi.nlm.nih.gov/pmc/utils/idconv/v1.0/"
                f"?ids={doi}&format=json"
            )
            r = requests.get(conv_url, timeout=8, headers=headers)
            if r.status_code == 200:
                records = r.json().get("records", [])
                pmcid = records[0].get("pmcid", "") if records else ""
        elif pmid:
            conv_url = (
                f"https://www.ncbi.nlm.nih.gov/pmc/utils/idconv/v1.0/"
                f"?ids={pmid}&idtype=pmid&format=json"
            )
            r = requests.get(conv_url, timeout=8, headers=headers)
            pmcid = ""
            if r.status_code == 200:
                records = r.json().get("records", [])
                pmcid = records[0].get("pmcid", "") if records else ""
        else:
            return None

        if pmcid:
            # Direct PMC PDF link
            pmcid_clean = pmcid.replace("PMC", "")
            return f"https://www.ncbi.nlm.nih.gov/pmc/articles/PMC{pmcid_clean}/pdf/"
    except Exception:
        pass
    return None


def find_open_access_pdf(doi: str = "", title: str = "", pmid: str = "") -> dict:
    """
    Tìm PDF Open Access from nhiều nguồn theo thứ tự ưu tiên:
    1. Unpaywall  2. Semantic Scholar  3. PubMed Central
    Trả về dict: {url, source, found}
    """
    if not HAS_REQUESTS:
        return {"url": None, "source": "", "found": False}

    # 1. Unpaywall
    if doi:
        url = fetch_unpaywall(doi)
        if url:
            return {"url": url, "source": "Unpaywall", "found": True}

    # 2. Semantic Scholar
    url = fetch_semantic_scholar_pdf(doi=doi, title=title)
    if url:
        return {"url": url, "source": "Semantic Scholar", "found": True}

    # 3. PubMed Central
    url = fetch_pmc_pdf(doi=doi, pmid=pmid)
    if url:
        return {"url": url, "source": "PubMed Central", "found": True}

    return {"url": None, "source": "", "found": False}


def _save_pdf_bytes(content: bytes, doi: str = "", filename_hint: str = "") -> dict:
    """
    Save bytes PDF ando storage and trích xuất metadata.
    Hàm nội bộ dùng chung cho cả tải from URL and xử lý bytes trực tiếp.
    Trả về dict: {file_path, pages, full_text, abstract_extracted, success, error}
    """
    # Validate this is a valid PDF
    if len(content) < 5000 or b"%PDF" not in content[:1024]:
        return {"success": False, "error": "Not a valid PDF file or file too small"}

    md5 = hashlib.md5(content).hexdigest()
    safe_name = re.sub(r'[^\w\-]', '_', (filename_hint or doi or md5[:8]))[:40]
    dest = STORAGE_DIR / f"{md5[:8]}_{safe_name}.pdf"

    with open(dest, "wb") as f:
        f.write(content)

    result: dict = {"success": True, "file_path": str(dest), "pages": 0,
                    "full_text": "", "abstract_extracted": ""}

    if HAS_PYMUPDF:
        try:
            # Truyền bytes trực tiếp — không đọc lại from disk
            pdf_data = extract_pdf_text_and_meta(content)
            result["pages"]              = pdf_data.get("pages", 0)
            result["full_text"]          = pdf_data.get("full_text", "")[:50000]
            result["abstract_extracted"] = pdf_data.get("abstract", "")
        except Exception as e:
            result["extract_error"] = str(e)

    return result


def download_and_save_pdf(pdf_url: str, doi: str = "", filename_hint: str = "") -> dict:
    """
    Tải PDF from URL and lưu ando storage.
    Trả về dict: {file_path, pages, full_text, abstract_extracted, success, error}
    """
    if not HAS_REQUESTS or not pdf_url:
        return {"success": False, "error": "requests not installed or URL empty"}
    try:
        r = requests.get(
            pdf_url, timeout=30,
            headers={"User-Agent": "Mozilla/5.0 SciKMS/1.0"},
            allow_redirects=True,
        )
        if r.status_code != 200:
            return {"success": False, "error": f"HTTP {r.status_code}"}
        content = r.content
        # Accept if header has pdf or magic bytes has %PDF
        if b"%PDF" not in content[:1024] and "pdf" not in r.headers.get("Content-Type","").lower():
            return {"success": False, "error": "Response is not PDF"}
        return _save_pdf_bytes(content, doi=doi, filename_hint=filename_hint)
    except requests.exceptions.Timeout:
        return {"success": False, "error": "Timeout when downloading PDF (30s)"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def save_pdf_bytes_direct(content: bytes, doi: str = "", filename_hint: str = "") -> dict:
    """Save PDF from bytes đã có sẵn (ví dụ: file upload) — không cần tải lại."""
    return _save_pdf_bytes(content, doi=doi, filename_hint=filename_hint)


def import_by_doi_only(doi: str, auto_download_pdf: bool = False) -> dict:
    """
    Tạo bản ghi paper from DOI mà không cần file PDF.
    Use Crossref for metadata, Unpaywall to find OA PDF.
    """
    doi = doi.strip().lstrip("https://doi.org/").lstrip("http://doi.org/")

    # Check duplicates DOI
    conn = get_db()
    exists = conn.execute("SELECT id,title FROM papers WHERE doi=?", (doi,)).fetchone()
    conn.close()
    if exists:
        return {"error": f"DOI already exists in library: «{exists['title']}»"}

    # Lấy metadata from Crossref
    meta = fetch_crossref(doi)
    if not meta or not meta.get("title"):
        return {"error": f"Metadata not found for DOI: {doi}. Please check DOI."}

    file_path = ""
    pages = 0
    full_text = f"{meta.get('title','')} {meta.get('abstract','')}"
    pdf_url_found = None
    pdf_source = ""

    # Thử tìm and tải PDF Open Access from nhiều nguồn
    if auto_download_pdf:
        oa = find_open_access_pdf(doi=doi, title=meta.get("title", ""))
        if oa["found"]:
            pdf_url_found = oa["url"]
            pdf_source    = oa["source"]
            dl = download_and_save_pdf(pdf_url_found, doi=doi,
                                       filename_hint=re.sub(r'[^\w]', '_', meta.get("title",""))[:30])
            if dl["success"]:
                file_path = dl["file_path"]
                pages     = dl.get("pages", 0)
                if dl.get("full_text"):
                    full_text = dl["full_text"]
                if not meta.get("abstract") and dl.get("abstract_extracted"):
                    meta["abstract"] = dl["abstract_extracted"]

    # Calculate MD5: use file if available, otherwise use DOI
    if file_path and os.path.exists(file_path):
        with open(file_path, "rb") as _f:
            md5 = hashlib.md5(_f.read()).hexdigest()
    else:
        md5 = hashlib.md5(doi.encode()).hexdigest()

    tags = auto_tag(full_text[:5000], meta.get("keywords", ""), meta.get("abstract", ""))
    keywords = meta.get("keywords", "")

    paper = {
        "md5": md5,
        "original_filename": f"(DOI import: {doi})",
        "title": meta.get("title", ""),
        "authors": meta.get("authors", ""),
        "year": meta.get("year") or datetime.now().year,
        "journal": meta.get("journal", ""),
        "doi": doi,
        "abstract": meta.get("abstract", ""),
        "keywords": keywords,
        "full_text": full_text,
        "tags": json.dumps(tags),
        "status": "unread",
        "starred": 0,
        "pages": pages,
        "added_at": datetime.now().strftime("%Y-%m-%d"),
        "file_path": file_path,
        "notes": "",
        "highlights": "[]",
        "project": "",
    }
    paper["renamed_filename"] = build_renamed_filename(paper)

    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO papers
            (md5,original_filename,renamed_filename,title,authors,year,journal,doi,
             abstract,keywords,full_text,tags,notes,highlights,status,starred,pages,added_at,file_path,project)
            VALUES (:md5,:original_filename,:renamed_filename,:title,:authors,:year,:journal,:doi,
                    :abstract,:keywords,:full_text,:tags,:notes,:highlights,:status,:starred,:pages,:added_at,:file_path,:project)
        """, paper)
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return {"error": "Paper already exists in library (duplicate)."}
    finally:
        conn.close()

    paper["_pdf_url"] = pdf_url_found or ""
    paper["_pdf_downloaded"] = bool(file_path)
    paper["_pdf_source"] = pdf_source
    return paper


def generate_citation(paper: dict, fmt: str = "vancouver") -> str:
    """
    Generate standard citation string.
    fmt: 'vancouver' | 'apa' | 'chicago'
    """
    authors_raw = paper.get("authors") or ""
    author_list = [a.strip() for a in authors_raw.split(";") if a.strip()]
    title   = paper.get("title") or ""
    journal = paper.get("journal") or ""
    year    = paper.get("year") or ""
    doi     = paper.get("doi") or ""
    pages   = paper.get("pages") or ""
    doi_str = f" doi:{doi}" if doi else ""

    def fmt_author_vancouver(a: str) -> str:
        parts = [x.strip() for x in a.split(",")]
        if len(parts) >= 2:
            last, first = parts[0], parts[1]
            initials = "".join(w[0].upper() for w in first.split() if w)
            return f"{last} {initials}"
        return a

    def fmt_author_apa(a: str) -> str:
        parts = [x.strip() for x in a.split(",")]
        if len(parts) >= 2:
            last, first = parts[0], parts[1]
            initials = ". ".join(w[0].upper() for w in first.split() if w) + "."
            return f"{last}, {initials}"
        return a

    if fmt == "vancouver":
        if len(author_list) <= 6:
            auth_str = ", ".join(fmt_author_vancouver(a) for a in author_list)
        else:
            auth_str = ", ".join(fmt_author_vancouver(a) for a in author_list[:6]) + ", et al"
        p_str = f":{pages}" if pages else ""
        return f"{auth_str}. {title}. {journal}. {year}{p_str}.{doi_str}"

    elif fmt == "apa":
        if len(author_list) <= 7:
            auth_str = ", ".join(fmt_author_apa(a) for a in author_list)
        else:
            auth_str = ", ".join(fmt_author_apa(a) for a in author_list[:6]) + ", ... " + fmt_author_apa(author_list[-1])
        doi_apa = f" https://doi.org/{doi}" if doi else ""
        return f"{auth_str} ({year}). {title}. {journal}.{doi_apa}"

    elif fmt == "chicago":
        if len(author_list) == 1:
            auth_str = author_list[0]
        elif len(author_list) <= 3:
            auth_str = ", ".join(author_list[:-1]) + ", and " + author_list[-1]
        else:
            auth_str = author_list[0] + " et al."
        return f"{auth_str}. \"{title}.\" {journal} ({year}).{doi_str}"

    return ""


def check_smart_duplicate(title: str, doi: str) -> dict | None:
    """
    Smart duplicate check by DOI or title similarity.
    Returns existing paper if duplicated, None otherwise.
    Use context manager to ensure connection is always closed.
    """
    with get_db() as conn:
        conn.row_factory = sqlite3.Row
        # Check DOI first
        if doi:
            row = conn.execute(
                "SELECT id,title,doi FROM papers WHERE doi=? AND doi!=''", (doi,)
            ).fetchone()
            if row:
                return dict(row)
        # Check similar title (Jaccard similarity)
        if title and len(title) > 10:
            norm_title = re.sub(r'\s+', ' ', title.lower().strip())
            rows = conn.execute(
                "SELECT id,title,doi FROM papers WHERE length(title)>10"
            ).fetchall()
            for r in rows:
                db_title = re.sub(r'\s+', ' ', (r["title"] or "").lower().strip())
                if db_title and _title_similarity(norm_title, db_title) > 0.85:
                    return dict(r)
    return None


def _title_similarity(a: str, b: str) -> float:
    """Calculate simple similarity between 2 strings (Jaccard token similarity)."""
    set_a = set(a.split())
    set_b = set(b.split())
    if not set_a or not set_b:
        return 0.0
    intersection = len(set_a & set_b)
    union = len(set_a | set_b)
    return intersection / union if union else 0.0


def get_all_projects() -> list[str]:
    """Lấy danh sách projects from DB."""
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT project FROM papers WHERE project IS NOT NULL AND project != '' ORDER BY project").fetchall()
    conn.close()
    return [r["project"] for r in rows]


def extract_pdf_text_and_meta(pdf_bytes: bytes) -> dict:
    """Trích xuất text + metadata from PDF bằng PyMuPDF."""
    if not HAS_PYMUPDF:
        return {"full_text": "", "pages": 0, "abstract": "", "title": "", "author": "", "keywords": ""}

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages = len(doc)
    full_text = ""
    for i, page in enumerate(doc):
        full_text += page.get_text()
        if i > 20:   # Limit first 20 pages to save memory
            break

    meta = doc.metadata or {}

    # Rút keywords from metadata hoặc from text
    kw_meta = meta.get("keywords") or ""

    # Rút author from PDF metadata
    pdf_author = meta.get("author") or ""

    # If not in metadata, try regex on first segment
    if not pdf_author:
        first_page = doc[0].get_text() if pages > 0 else ""
        # Try pattern: "Author(s): ..." or line after title before abstract
        author_match = re.search(
            r'(?:authors?|author)[:\s]+([^\n]{5,200})',
            first_page, re.IGNORECASE
        )
        if author_match:
            pdf_author = author_match.group(1).strip()

    # ── Abstract extraction - Dark optimized to avoid ReDoS (Freezing) ─────────────────
    abstract = ""
    first_2000 = full_text[:8000]
    text_lower = first_2000.lower()
    
    # Find start position of Abstract
    abs_start = text_lower.find("abstract")
    if abs_start == -1:
        abs_start = text_lower.find("abstract")
        
    if abs_start != -1:
        # Tìm vị trí kết thúc (dựa ando các tiêu đề theo sau phổ biến)
        end_markers = ["keywords", "key words", "1. introduction", "\nintroduction", "\nbackground"]
        abs_end = len(first_2000)
        
        # Find nearest marker appearing AFTER abstract
        for marker in end_markers:
            idx = text_lower.find(marker, abs_start + 10) # Skip first 10 chars to avoid matching itself
            if idx != -1 and idx < abs_end:
                abs_end = idx
                
        # Extract if length is reasonable (avoid taking whole paper)
        if 50 < (abs_end - abs_start) < 4000:
            # Get original text (to preserve case)
            raw_abs = first_2000[abs_start:abs_end]
            # Remove "Abstract" at the beginning
            raw_abs = re.sub(r'^(?i)(abstract|abstract)[\s:\.\n]*', '', raw_abs).strip()
            # Clean whitespaces
            abstract = re.sub(r'\s{2,}', ' ', raw_abs.replace("\n", " "))
            
    # Limit reasonable length
    if len(abstract) > 2500:
        abstract = abstract[:2500] + "…"

    # Rút keywords from text nếu chưa có
    if not kw_meta:
        kw_match = re.search(
            r'(?:keywords?|key\s*words?)[:\s]+([^\n]{10,300})',
            full_text, re.IGNORECASE
        )
        kw_meta = kw_match.group(1).strip() if kw_match else ""

    return {
        "full_text": full_text[:50000],  # Limit 50k characters
        "pages": pages,
        "abstract": abstract,
        "title": meta.get("title") or "",
        "author": pdf_author,
        "keywords": kw_meta,
        "first_page_text": (doc[0].get_text() if pages > 0 else "")[:3000],
    }


def extract_meta_with_gemini(first_page_text: str) -> dict:
    """Dùng Gemini AI để trích xuất metadata from pages đầu papers."""
    if not HAS_REQUESTS:
        return {}
    GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
    if not GEMINI_API_KEY:
        return {}  # Fail silently if no API key
        
    MODEL_NAME = 'gemini-2.5-flash'
    
    try:
        prompt = f"""Extract metadata from this academic paper's first page text. Return ONLY a JSON object with these keys: title, authors, year, journal, abstract, keywords.
- authors: semicolon-separated list of full author names (e.g. "Nguyen Van A; Tran Thi B"). DO NOT include affiliations or degrees.
- year: 4-digit year as integer
- keywords: comma-separated list of actual author-provided keywords or highly relevant technical terms. DO NOT include dates, universities, affiliations, or journal names.
- If a field cannot be found, use empty string or null

First page text:
{first_page_text[:2500]}

Return only valid JSON, no explanation."""

        url = f"https://generativelanguage.googleapis.com/v1beta/models/{MODEL_NAME}:generateContent?key={GEMINI_API_KEY}"
        
        resp = requests.post(
            url,
            headers={"Content-Type": "application/json"},
            json={
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.1}
            },
            timeout=20,
        )
        if resp.status_code == 200:
            text = resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
            # Strip markdown fences if any
            text = re.sub(r'^```json\s*|^```\s*|```$', '', text, flags=re.MULTILINE).strip()
            return json.loads(text)
    except Exception as e:
        print(f"Gemini API Error: {e}")
    return {}


# ─── Default Tag dictionary (Plastic / Reconstructive Surgery) ──────────────
_DEFAULT_TAG_DICT = [
    "Photogrammetry", "3D Morphometrics", "Local Flaps", "Farkas anthropometry",
    "Rhinoplasty", "Osteotomy", "Bone Graft", "Microsurgery", "Perforator Flap",
    "Machine Learning", "Deep Learning", "Artificial Intelligence",
    "Augmented Reality", "Virtual Reality", "Surgical Planning",
    "Maxillofacial", "Craniofacial", "Cleft Palate", "Burn Reconstruction",
    "Scar Management", "Keloid", "POSAS", "Patient Reported Outcome",
    "Meta-analysis", "Systematic Review", "RCT", "Cohort Study",
    "Anthropometry", "Biometrics", "Finite Element Analysis",
    "Stereophotogrammetry", "CBCT", "CT Scan", "MRI",
    "Nasal Tip", "Nasal Dorsum", "Septoplasty",
    "Fat Grafting", "Liposuction", "Abdominoplasty",
    "Breast Reconstruction", "TRAM Flap", "DIEP Flap",
    # Internal medicine / general
    "Hypertension", "Diabetes", "Cardiovascular", "Oncology", "Pharmacology",
    "Clinical Trial", "Case Report", "Cross-sectional", "Prospective Study",
]


def get_custom_tag_dict() -> list[str]:
    """Đọc from điển tag tùy chỉnh from config. Fallback về dict mặc định."""
    cfg = read_config()
    user_dict = cfg.get("custom_tag_dict")
    if isinstance(user_dict, list) and len(user_dict) > 0:
        return user_dict
    return _DEFAULT_TAG_DICT


def save_custom_tag_dict(terms: list[str]):
    """Save from điển tag tùy chỉnh ando config."""
    cfg = read_config()
    cfg["custom_tag_dict"] = [t.strip() for t in terms if t.strip()]
    save_config(cfg)


def auto_tag(text: str, keywords: str, abstract: str) -> list[str]:
    """Tự động gắn tag from from điển chuyên ngành (có thể tùy chỉnh)."""
    tag_dict = get_custom_tag_dict()
    combined = f"{text} {keywords} {abstract}".lower()
    found = []
    for term in tag_dict:
        if term.lower() in combined:
            found.append(term)
    return found[:10]  # Max 10 tags


def heuristic_fallback_extract(first_page_text: str) -> dict:
    """
    Fallback logic: Analyze Header block of papers.
    Only use when basic Regex cannot find author/title.
    """
    # 1. Cắt lấy phần "Header" (from đầu đến trước chữ Abstract/Introduction)
    parts = re.split(r'(?i)\n\s*(abstract|abstract|introduction|background)\b', first_page_text, maxsplit=1)
    header_block = parts[0]
    
    lines = [line.strip() for line in header_block.split('\n') if line.strip()]
    
    # Tập from khóa nhận diện rác (thông tin xuất bản)
    junk_kws = [r'(?i)^vol\.', r'(?i)^issn', r'(?i)doi:', r'^http', r'(?i)copyright', r'(?i)received', r'(?i)accepted']
    
    # Tập from khóa nhận diện Cơ quan/Đơn vị công tác (Affiliation) - Hỗ trợ cả English & Việt
    affil_kws = [
        r'(?i)university', r'(?i)department', r'(?i)institute', r'(?i)hospital', r'(?i)clinic',
        r'(?i)school', r'(?i)college', r'(?i)center', r'(?i)laboratory',
        r'(?i)university', r'(?i)hospital', r'(?i)department ', r'(?i)institute ', r'@'
    ]
    
    clean_lines = [l for l in lines if not any(re.search(kw, l) for kw in junk_kws)]
    if not clean_lines:
        return {}

    title = ""
    author_start_idx = 0
    
    # 2. Tìm Title: Thường là dòng dài đầu tiên, không chứa from khóa cơ quan
    for i, line in enumerate(clean_lines):
        if any(re.search(kw, line) for kw in affil_kws):
            continue
        if len(line.split()) > 2:  # Title thường có nhiều hơn 2 from
            title = line
            author_start_idx = i + 1
            break
            
    # 3. Find Authors: Lines right below title, stop when hitting affiliation keyword
    author_lines = []
    for line in clean_lines[author_start_idx:]:
        # If hitting line containing Hospital, Department, School -> skip
        if any(re.search(kw, line) for kw in affil_kws):
            continue
        
        # Person names usually in lines not too long
        if len(line.split()) < 20:
            author_lines.append(line)
            
    authors_str = ""
    if author_lines:
        raw_authors = " ".join(author_lines)
        
        # Clean up superscript numbers, footnote letters (e.g., Ball¹, Nguyen^a, Tran*)
        clean_authors = re.sub(r'[\d¹²³⁴⁵⁶⁷⁸⁹⁰\*†‡§]', '', raw_authors)
        # Remove letters a, b, c next to commas (e.g., Smith a, Jones b)
        clean_authors = re.sub(r'\s+[a-z]\s*(?=[,\s])', '', clean_authors)
        
        # Split names by comma, "and", "&"
        splits = re.split(r'(?i)\s+and\s+|&|,', clean_authors)
        valid_authors = [a.strip() for a in splits if len(a.strip()) > 2]
        authors_str = "; ".join(valid_authors)
        
    return {"title": title.title() if title else "", "authors": authors_str}


def extract_vn_thesis_meta(first_page_text: str) -> dict:
    """
    Powerful extraction specialized for Vietnamese Thesis covers.
    Bắt chính xác Title nhiều dòng, dọn dẹp sạch sẽ học hàm/học vị, and lấy được Year chuẩn xác.
    """
    lines = [l.strip() for l in first_page_text.split('\n') if l.strip()]
    
    author = ""
    title = ""
    year = None
    
    # ==========================================
    # 1. GET DEFENSE/PUBLICATION YEAR (UPGRADED)
    # ==========================================
    # Priority 1: Find clear patterns at end of pages like "Year 2020", "Hà Nội, 2020", "Huế - 2020"
    explicit_year_match = re.search(r'(?i)(?:năm|hà nội|huế|hồ chí minh|đà nẵng|cần thơ|thái nguyên|hải phòng)[\s\,\-\.]+([12]\d{3})\b', first_page_text)
    
    if explicit_year_match:
        year = int(explicit_year_match.group(1))
    else:
        # Ưu tiên 2: Tìm tất cả các năm (from 1950 - 2030) and lấy năm CUỐI CÙNG xuất hiện trên pages bìa
        all_years = re.findall(r'\b(19[5-9]\d|20[0-3]\d)\b', first_page_text)
        if all_years:
            year = int(all_years[-1]) # Get last element because defense year is always at end of pages

    # ==========================================
    # 2. FIND AUTHOR AND CLEAN DEGREES
    # ==========================================
    author_labels = [
        r"phd student", r"student", r"student", r"author", 
        r"executor", r"họ and tên(?:\s*(?:ncs|student|student))?", r"full name"
    ]
    label_pattern = r'(?i)^(?:' + '|'.join(author_labels) + r')\s*[:\-]*\s*(.*)'
    
    for i, line in enumerate(lines):
        m = re.search(label_pattern, line)
        if m:
            potential_name = m.group(1).strip()
            # Nếu nhãn ở 1 dòng and tên rớt xuống dòng dưới
            if not potential_name and i + 1 < len(lines):
                potential_name = lines[i+1].strip()
            author = potential_name
            break

    # Thoroughly clean combined degree abbreviations (e.g., "PGS.TS.BS.", "ThS.BS")
    if author:
        prefix_pattern = r'(?i)^(?:GS|PGS|TS|ThS|Ths|BS|Bs|CKI|CKII|NCS|CN|\.)[\s\.\,\-]*'
        # Use loop to clean string if multiple consecutive degrees exist
        while re.match(prefix_pattern, author):
            author = re.sub(prefix_pattern, '', author).strip()

    # ==========================================
    # 3. FIND THESIS/PAPER TITLE
    # ==========================================
    junk_kws = [
        "MINISTRY OF EDUCATION", "MINISTRY OF HEALTH", "UNIVERSITY", "SCHOOL", "INSTITUTE", 
        "THESIS", "DISSERTATION", "THESIS", "REPORT", "MAJOR", 
        "CODE", "SUPERVISOR", "NĂM", "HUẾ", "HÀ NỘI", "HỒ CHÍ MINH",
        "CITY", "SPECIALTY", "OFFICIAL", "MINISTRY OF DEFENSE"
    ]
    
    title_lines = []
    found_de_tai = False
    
    # Scenario 1: Capture by "TOPIC" anchor
    for i, line in enumerate(lines):
        if re.search(r'(?i)^(?:tên\s*)?đề tài\s*[:\-]', line):
            found_de_tai = True
            content = re.sub(r'(?i)^(?:tên\s*)?đề tài\s*[:\-]\s*', '', line).strip()
            if content: 
                title_lines.append(content)
                
            # Scan below lines to concatenate title (max 6 lines)
            for j in range(i+1, min(i+6, len(lines))):
                next_line = lines[j]
                if any(k in next_line.upper() for k in junk_kws) or re.search(r'(?i)phd student|student|họ and tên|supervisor', next_line):
                    break
                title_lines.append(next_line)
            break
            
    # Scenario 2: Without TOPIC word, find longest UPPERCASE block
    if not found_de_tai:
        current_block = []
        best_block = []
        for line in lines:
            upper_line = line.upper()
            is_junk = any(k in upper_line for k in junk_kws)
            # Dòng tiêu đề thường viết hoa, dài hơn 2 from and không chứa rác
            if line.isupper() and not is_junk and len(line.split()) >= 2:
                current_block.append(line)
            else:
                if len(current_block) > len(best_block):
                    best_block = current_block
                current_block = []
        if len(current_block) > len(best_block):
            best_block = current_block
            
        title_lines = best_block

    if title_lines:
        title = re.sub(r'\s+', ' ', " ".join(title_lines))

    return {
        "title": title.title() if title else "", 
        "authors": author.title() if author else "",
        "year": year
    }


def is_garbage_title(t: str) -> bool:
    """Kiểm tra xem title lấy from pdf properties có phải là rác không"""
    if not t or len(t) < 5: return True
    if "_" in t and " " not in t: return True
    return False


def is_garbage_author(a: str) -> bool:
    """Kiểm tra xem author lấy from pdf properties có phải là rác không"""
    if not a or len(a) < 3: return True
    if " " not in a and a.islower(): return True
    return False


def process_uploaded_pdf(uploaded_file) -> dict | None:
    """Pipeline for processing uploaded PDF files."""
    file_bytes = uploaded_file.read()
    md5 = md5_of_bytes(file_bytes)

    # Check duplicate (MD5 hash)
    conn = get_db()
    exists = conn.execute("SELECT id FROM papers WHERE md5=?", (md5,)).fetchone()
    conn.close()
    if exists:
        return {"error": f"File already exists in library (MD5: {md5[:8]}…)"}

    # Trích xuất from PDF
    pdf_data = extract_pdf_text_and_meta(file_bytes)
    full_text = pdf_data["full_text"]

    # Step 1: Find DOI
    doi = extract_doi(full_text) or extract_doi(uploaded_file.name)

    meta = {}
    # Step 2: Call Crossref if DOI exists
    if doi:
        meta = fetch_crossref(doi)

    # Step 3: Use PDF metadata if insufficient (With junk filter)
    raw_pdf_title = pdf_data.get("title", "")
    raw_pdf_author = pdf_data.get("author", "")

    if not meta.get("title") and not is_garbage_title(raw_pdf_title):
        meta["title"] = raw_pdf_title
        
    if not meta.get("abstract") and pdf_data.get("abstract"):
        meta["abstract"] = pdf_data["abstract"]
        
    # Ưu tiên author from PDF metadata nhưng phải qua kiểm duyệt
    if not meta.get("authors") and not is_garbage_author(raw_pdf_author):
        meta["authors"] = raw_pdf_author

    # ==========================================
    # STEP 3.5: FALLBACK HEURISTIC
    # Only run when old Regex fails to capture Title or Authors
    # ==========================================
    if not meta.get("title") or not meta.get("authors"):
        first_page = pdf_data.get("first_page_text", "")
        fallback_meta = heuristic_fallback_extract(first_page)
        
        # Chỉ bổ sung ando các trường đang trống
        if not meta.get("title") and fallback_meta.get("title"):
            meta["title"] = fallback_meta["title"]
            
        if not meta.get("authors") and fallback_meta.get("authors"):
            meta["authors"] = fallback_meta["authors"]

    # ==========================================
    # BƯỚC 3.6: LOGIC CHUYÊN DỤNG CHO THESIS/TÀI LIỆU TIẾNG VIỆT
    # ==========================================
    if not meta.get("title") or not meta.get("authors"):
        first_page = pdf_data.get("first_page_text", "")
        # Identify characteristics of VN docs (Thesis, Health, Education...)
        if re.search(r'(?i)(luận án|luận văn|khoá luận|bộ giáo dục|bộ y tế|trường đại học|phd student|chuyên khoa|đề tài)', first_page):
            vn_meta = extract_vn_thesis_meta(first_page)
            
            # Overwrite or supplement information
            if not meta.get("title") and vn_meta.get("title"):
                meta["title"] = vn_meta["title"]
            if not meta.get("authors") and vn_meta.get("authors"):
                meta["authors"] = vn_meta["authors"]
            if not meta.get("year") and vn_meta.get("year"):
                meta["year"] = vn_meta["year"]

    # Step 4: Gemini AI fallback - when title or authors are missing
    if HAS_REQUESTS and (not meta.get("title") or not meta.get("authors")):
        first_page = pdf_data.get("first_page_text") or pdf_data.get("full_text", "")[:3000]
        ai_meta = extract_meta_with_gemini(first_page)
        if ai_meta:
            if not meta.get("title") and ai_meta.get("title"):
                meta["title"] = ai_meta["title"]
            if not meta.get("authors") and ai_meta.get("authors"):
                meta["authors"] = ai_meta["authors"]
            if not meta.get("year") and ai_meta.get("year"):
                meta["year"] = ai_meta["year"]
            if not meta.get("journal") and ai_meta.get("journal"):
                meta["journal"] = ai_meta["journal"]
            if not meta.get("abstract") and ai_meta.get("abstract"):
                meta["abstract"] = ai_meta["abstract"]
            if not meta.get("keywords") and ai_meta.get("keywords"):
                meta["keywords"] = ai_meta["keywords"]

    # Bước 5: Fallback thông minh from tên file
    if not meta.get("title"):
        meta["title"] = (
            Path(uploaded_file.name).stem
            .replace("_", " ").replace("-", " ").title()
        )
    if not meta.get("doi"):
        meta["doi"] = doi or ""

    # ─── SMART DUPLICATE CHECK (DOI + Title) ────────────────────────
    smart_dup = check_smart_duplicate(meta.get("title", ""), meta.get("doi", ""))
    if smart_dup:
        match_reason = f"DOI: {smart_dup.get('doi')}" if smart_dup.get("doi") == meta.get("doi") else f"Similar title"
        return {
            "error": f"⚠️ Detected papers trùng lặp ({match_reason}): «{smart_dup.get('title','')[:80]}»",
            "_is_smart_dup": True,
            "_dup_id": smart_dup.get("id"),
        }

    # Tự động gắn tags from from điển chuyên ngành
    tags = auto_tag(full_text[:5000], meta.get("keywords", ""), meta.get("abstract", ""))
    
    # Lấy keywords from AI hoặc PDF metadata (ưu tiên AI)
    final_keywords = meta.get("keywords") or pdf_data.get("keywords") or ""
    
    if final_keywords:
        # Bộ lọc rác: Chặn các from khóa chứa tên cơ quan, học vị hoặc ngày tháng
        junk_words = [
            "university", "institute", "department", "hospital", "school", "college", 
            "society", "research", "academy", "clinic", "center",
            "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"
        ]
        
        for kw in final_keywords.split(","):
            kw = kw.strip()
            kw_lower = kw.lower()
            
            # Check exclusion conditions: 
            # 1. Chứa from rác
            # 2. Is date string (e.g. 09 Jun 2020)
            # 3. Too long (>45 chars is usually a sentence not keyword)
            is_junk = (
                any(junk in kw_lower for junk in junk_words) or 
                re.search(r'\d{1,2}\s+[a-z]{3}\s+\d{4}', kw_lower) or 
                len(kw) > 45
            )
            
            if kw and not is_junk and kw not in tags:
                tags.append(kw)

    # Save file ando storage/
    safe_name = re.sub(r'[^\w\-.]', '_', uploaded_file.name)
    dest = STORAGE_DIR / f"{md5[:8]}_{safe_name}"
    with open(dest, "wb") as f:
        f.write(file_bytes)

    paper = {
        "md5": md5,
        "original_filename": uploaded_file.name,
        "title": meta.get("title", ""),
        "authors": meta.get("authors", ""),
        "year": meta.get("year") or datetime.now().year,
        "journal": meta.get("journal", ""),
        "doi": meta.get("doi", ""),
        "abstract": meta.get("abstract", ""),
        "keywords": final_keywords,
        "full_text": full_text,
        "tags": json.dumps(tags),
        "status": "unread",
        "starred": 0,
        "pages": pdf_data["pages"],
        "added_at": datetime.now().strftime("%Y-%m-%d"),
        "file_path": str(dest),
        "notes": "",
        "highlights": "[]",
        "project": "",
    }
    paper["renamed_filename"] = build_renamed_filename(paper)

    # Ghi ando DB
    conn = get_db()
    conn.execute("""
        INSERT INTO papers
        (md5,original_filename,renamed_filename,title,authors,year,journal,doi,
         abstract,keywords,full_text,tags,notes,highlights,status,starred,pages,added_at,file_path,project)
        VALUES (:md5,:original_filename,:renamed_filename,:title,:authors,:year,:journal,:doi,
                :abstract,:keywords,:full_text,:tags,:notes,:highlights,:status,:starred,:pages,:added_at,:file_path,:project)
    """, paper)
    conn.commit()
    conn.close()
    return paper


# Columns needed for list view - exclude full_text to reduce RAM
_LIST_COLUMNS = (
    "id,md5,original_filename,renamed_filename,title,authors,year,journal,"
    "doi,abstract,keywords,tags,notes,highlights,status,starred,pages,"
    "added_at,file_path,project,reading_position"
)

@st.cache_data(ttl=30)
def get_all_papers(order: str = "added_at DESC") -> list[dict]:
    """Load all papers but exclude full_text to save RAM.
    Used for Library, Search, Sidebar stats. TTL=30s."""
    _VALID_ORDERS = {
        "added_at DESC", "year DESC", "year ASC",
        "title ASC", "title DESC", "authors ASC", "authors DESC",
    }
    safe_order = order if order in _VALID_ORDERS else "added_at DESC"
    conn = get_db()
    try:
        rows = conn.execute(
            f"SELECT {_LIST_COLUMNS} FROM papers ORDER BY {safe_order}"
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


def get_paper_full(paper_id: int) -> dict | None:
    """Load full paper including full_text - only used when reading."""
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM papers WHERE id=?", (paper_id,)).fetchone()
    finally:
        conn.close()
    return dict(row) if row else None


def get_papers_count() -> int:
    """Count number of papers - avoid loading data."""
    conn = get_db()
    try:
        return conn.execute("SELECT COUNT(*) FROM papers").fetchone()[0]
    finally:
        conn.close()


@st.cache_data(ttl=30)
def get_papers_page(order: str = "added_at DESC", page: int = 0, page_size: int = 20) -> list[dict]:
    """Load papers by page - used for Library with large collection."""
    _VALID_ORDERS = {
        "added_at DESC", "year DESC", "year ASC",
        "title ASC", "title DESC", "authors ASC", "authors DESC",
    }
    safe_order = order if order in _VALID_ORDERS else "added_at DESC"
    offset = page * page_size
    conn = get_db()
    try:
        rows = conn.execute(
            f"SELECT {_LIST_COLUMNS} FROM papers ORDER BY {safe_order} LIMIT ? OFFSET ?",
            (page_size, offset)
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


# Whitelist valid columns in papers table - protect against SQL injection
_PAPER_ALLOWED_COLUMNS = frozenset({
    "title", "authors", "year", "journal", "doi", "abstract", "keywords",
    "full_text", "tags", "notes", "highlights", "status", "starred", "pages",
    "added_at", "file_path", "project", "renamed_filename", "original_filename",
    "reading_position",
})

def update_paper(paper_id: int, fields: dict):
    """Safely update paper with explicit column whitelist."""
    if not fields:
        return
    # Filter to keep only valid columns - prevent SQL injection via column name
    safe_fields = {k: v for k, v in fields.items() if k in _PAPER_ALLOWED_COLUMNS}
    if not safe_fields:
        return
    set_clause = ", ".join(f"{k}=?" for k in safe_fields)
    vals = list(safe_fields.values()) + [paper_id]
    conn = get_db()
    try:
        conn.execute(f"UPDATE papers SET {set_clause} WHERE id=?", vals)
        conn.commit()
    finally:
        conn.close()


def delete_paper(paper_id: int):
    conn = get_db()
    row = conn.execute("SELECT file_path FROM papers WHERE id=?", (paper_id,)).fetchone()
    if row and row["file_path"] and os.path.exists(row["file_path"]):
        try:
            os.remove(row["file_path"])
        except Exception:
            pass
    conn.execute("DELETE FROM papers WHERE id=?", (paper_id,))
    conn.commit()
    conn.close()


def delete_all_papers():
    """Delete all papers and PDF files."""
    conn = get_db()
    rows = conn.execute("SELECT file_path FROM papers").fetchall()
    for row in rows:
        fp = row["file_path"]
        if fp and os.path.exists(fp):
            try:
                os.remove(fp)
            except Exception:
                pass
    conn.execute("DELETE FROM papers")
    conn.commit()
    conn.close()


def fts_search(query: str) -> list[dict]:
    """Search full-text với SQLite FTS5. Fallback về LIKE nếu lỗi.
    Returns list of dicts without full_text to save RAM."""
    if not query.strip():
        return get_all_papers()
    # Sanitize: keep alphanumeric + whitespace, remove special FTS5 chars
    safe_q = re.sub(r'[^\w\s]', ' ', query).strip()
    # Add * for prefix-match (e.g. "rhinop" -> "rhinoplasty")
    fts_q = " OR ".join(f'"{w}"*' for w in safe_q.split() if w) if safe_q else ""

    conn = get_db()
    try:
        if fts_q:
            rows = conn.execute(f"""
                SELECT {_LIST_COLUMNS}, fts.rank
                FROM papers p
                JOIN papers_fts fts ON p.id = fts.rowid
                WHERE papers_fts MATCH ?
                ORDER BY fts.rank
            """, (fts_q,)).fetchall()
        else:
            rows = []
    except Exception:
        rows = []
        # Fallback: LIKE search on important columns (not full_text)
        like = f"%{query}%"
        try:
            rows = conn.execute(f"""
                SELECT {_LIST_COLUMNS} FROM papers
                WHERE title LIKE ? OR authors LIKE ?
                   OR abstract LIKE ? OR keywords LIKE ?
                ORDER BY added_at DESC
            """, (like, like, like, like)).fetchall()
        except Exception:
            pass
    finally:
        conn.close()
    return [dict(r) for r in rows]


def highlight_text(text: str, query: str, max_len: int = 300) -> str:
    """Tô màu from khóa trong đoạn text, rút gọn cho preview."""
    if not query or not text:
        return text[:max_len] + ("…" if len(text) > max_len else "")
    idx = text.lower().find(query.lower())
    if idx == -1:
        return text[:max_len] + ("…" if len(text) > max_len else "")
    start = max(0, idx - 80)
    end = min(len(text), idx + len(query) + 80)
    snippet = ("…" if start > 0 else "") + text[start:end] + ("…" if end < len(text) else "")
    highlighted = re.sub(
        re.escape(query), f'<span class="search-match">{query}</span>', snippet, flags=re.IGNORECASE
    )
    return highlighted


# ══════════════════════════════════════════════════════════════════════════════
#  XUẤT FILE
# ══════════════════════════════════════════════════════════════════════════════

def export_ris(papers: list[dict]) -> str:
    lines = []
    for p in papers:
        lines += [
            "TY  - JOUR",
            f"TI  - {p.get('title','')}",
            f"PY  - {p.get('year','')}",
            f"JO  - {p.get('journal','')}",
            f"DO  - {p.get('doi','')}",
            f"AB  - {p.get('abstract','')}",
        ]
        for a in (p.get("authors") or "").split(";"):
            if a.strip():
                lines.append(f"AU  - {a.strip()}")
        kws = p.get("keywords") or ""
        for kw in kws.split(","):
            if kw.strip():
                lines.append(f"KW  - {kw.strip()}")
        lines.append("ER  -\n")
    return "\n".join(lines)


def export_bib(papers: list[dict]) -> str:
    entries = []
    for p in papers:
        author_list = (p.get("authors") or "").split(";")
        first_author = author_list[0].split(",")[0].strip() if author_list else "unknown"
        key = f"{slugify(first_author, 15)}{p.get('year','')}"
        entries.append(
            f"@article{{{key},\n"
            f"  title   = {{{p.get('title','')}}},\n"
            f"  author  = {{{' and '.join(a.strip() for a in author_list)}}},\n"
            f"  journal = {{{p.get('journal','')}}},\n"
            f"  year    = {{{p.get('year','')}}},\n"
            f"  doi     = {{{p.get('doi','')}}}\n}}"
        )
    return "\n\n".join(entries)


def export_excel(papers: list[dict]) -> bytes:
    """Xuất bảng abstract ra Excel với định dạng đẹp."""
    rows = []
    for p in papers:
        tags = json.loads(p.get("tags") or "[]")
        rows.append({
            "Original File Name": p.get("original_filename", ""),
            "Renamed File": p.get("renamed_filename", ""),
            "Year": p.get("year", ""),
            "Authors": p.get("authors", ""),
            "Title": p.get("title", ""),
            "Journals": p.get("journal", ""),
            "DOI": p.get("doi", ""),
            "Keywords": p.get("keywords", ""),
            "Tags": ", ".join(tags),
            "Abstract": (p.get("abstract") or "")[:500],
            "Status": p.get("status", ""),
            "Notes": p.get("notes", ""),
            "Date Added": p.get("added_at", ""),
        })
    df = pd.DataFrame(rows)

    if HAS_OPENPYXL:
        wb = Workbook()
        ws = wb.active
        ws.title = "Y văn tổng hợp"

        # Header colors
        GOLD = PatternFill("solid", fgColor="C9A84C")
        DARK = PatternFill("solid", fgColor="1A2030")
        thin = Border(
            left=Side(style='thin', color='2D3748'),
            right=Side(style='thin', color='2D3748'),
            top=Side(style='thin', color='2D3748'),
            bottom=Side(style='thin', color='2D3748')
        )

        # Header
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="0E1117", name="Calibri", size=11)
            cell.fill = GOLD
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin

        # Data
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            fill = PatternFill("solid", fgColor="141923") if r_idx % 2 == 0 else PatternFill("solid", fgColor="1A2030")
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=str(val) if val is not None else "")
                cell.font = Font(color="E2E8F0", name="Calibri", size=10)
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin

        # Auto column width
        col_widths = [20, 28, 8, 30, 50, 30, 25, 30, 25, 60, 12, 30, 14]
        for i, w in enumerate(col_widths[:len(df.columns)], 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        ws.row_dimensions[1].height = 28

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    else:
        # Fallback: CSV
        return df.to_csv(index=False).encode("utf-8-sig")


# ══════════════════════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
if "active_tab" not in st.session_state:
    st.session_state.active_tab = "📚 Library"
if "selected_paper_id" not in st.session_state:
    st.session_state.selected_paper_id = None
if "edit_mode" not in st.session_state:
    st.session_state.edit_mode = {}
if "search_query" not in st.session_state:
    st.session_state.search_query = ""
if "upload_results" not in st.session_state:
    st.session_state.upload_results = []
if "confirm_delall_lib" not in st.session_state:
    st.session_state.confirm_delall_lib = False
if "filter_project" not in st.session_state:
    st.session_state.filter_project = "All"


# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
        <div style="padding: 8px 0 24px">
            <div style="font-family:'Lora',serif; font-size:24px; font-weight:700; color:#c9a84c; line-height:1.2">
                📚 SciKMS
            </div>
            <div style="font-family:'IBM Plex Mono',monospace; font-size:10px; color:#94a3b8; letter-spacing:0.12em; margin-top:6px">
                SCIENTIFIC KNOWLEDGE MANAGEMENT
            </div>
            <div style="font-family:'IBM Plex Mono',monospace; font-size:10px; color:#94a3b8; letter-spacing:0.05em; margin-top:6px">
                Created by Dr. Hieuduy
            </div>
        </div>
    """, unsafe_allow_html=True)

    all_papers = get_all_papers()
    n_total  = len(all_papers)
    n_read   = sum(1 for p in all_papers if p["status"] == "read")
    n_rdg    = sum(1 for p in all_papers if p["status"] == "reading")
    n_star   = sum(1 for p in all_papers if p["starred"])
    n_proj   = len(get_all_projects())

    # Small stats
    st.markdown(f"""
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:24px">
            <div class="metric-box"><div class="metric-val">{n_total}</div><div class="metric-label">Papers</div></div>
            <div class="metric-box"><div class="metric-val">{n_read}</div><div class="metric-label">Read</div></div>
            <div class="metric-box"><div class="metric-val">{n_rdg}</div><div class="metric-label">Reading</div></div>
            <div class="metric-box"><div class="metric-val">{n_star}</div><div class="metric-label">★ Star</div></div>
            <div class="metric-box" style="grid-column:span 2"><div class="metric-val" style="font-size:20px">{n_proj}</div><div class="metric-label">🗂 Projects</div></div>
        </div>
    """, unsafe_allow_html=True)

    st.markdown('<div style="font-size:11px;font-weight:600;color:#94a3b8;text-transform:uppercase;letter-spacing:0.12em;margin-bottom:10px">Navigation</div>', unsafe_allow_html=True)

    tabs = [
        ("📚 Library",    "nav_library"),
        ("⬆️ Import",  "nav_import"),
        ("🔍 Search",   "nav_search"),
        ("✏️ Rename ","nav_rename"),
        ("📊 Summary","nav_summary"),
        ("📤 Export","nav_export"),
        ("⚙️ Settings",   "nav_settings"),
    ]

    # Map label → active_tab key (single-space version)
    tab_labels = [t[0] for t in tabs]
    tab_keys   = [t[1] for t in tabs]

    for label, key in tabs:
        active = st.session_state.active_tab == label
        if st.button(
            label,
            key=key,
            use_container_width=True,
            type="primary" if active else "secondary",
        ):
            st.session_state.active_tab = label
            st.rerun()

    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)

    # Quick filters
    st.markdown("**Filter**")
    filter_status = st.selectbox(
        "Status",
        ["All", "Unread", "Reading", "Read"],
        key="filter_status",
    )
    filter_starred = st.checkbox("⭐ Starred papers only", key="filter_starred")

    # ─── Project Filter ──────────────────────────────────────────────────────
    projects = get_all_projects()
    project_options = ["All"] + projects
    if len(projects) > 0:
        st.markdown("**🗂 Project**")
        filter_project = st.selectbox(
            "Project",
            project_options,
            key="filter_project",
            label_visibility="collapsed",
        )
    else:
        st.session_state.filter_project = "All"

    _sort_opts = ["Recently added", "Year (newest)", "Title A→Z", "Authors A→Z"]
    sort_by = st.selectbox("Sort by", _sort_opts, key="sort_by")

    if not HAS_PYMUPDF:
        st.markdown("""
            <div class="info-tip" style="margin-top:16px">
            ⚠️ <b>PyMuPDF not installed</b><br>
            <code>pip install pymupdf</code><br>
            PDF extraction features will be limited.
            </div>
        """, unsafe_allow_html=True)
    if not HAS_REQUESTS:
        st.markdown("""
            <div class="info-tip" style="margin-top:8px">
            ⚠️ <b>requests not installed</b><br>
            <code>pip install requests</code><br>
            Online access required for DOI lookup via Crossref.
            </div>
        """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  PAPER CARD DISPLAY FUNCTION
# ══════════════════════════════════════════════════════════════════════════════

def cb_quick_read(pid):
    st.session_state.selected_paper_id = pid
    st.session_state.active_tab = "📚 Library"

def cb_toggle_edit(pid):
    st.session_state.edit_mode[pid] = not st.session_state.edit_mode.get(pid, False)

def cb_toggle_star(pid, starred):
    update_paper(pid, {"starred": 0 if starred else 1})

def cb_toggle_status(pid, current_status):
    new_status = {"unread": "reading", "reading": "read", "read": "unread"}[current_status]
    update_paper(pid, {"status": new_status})

def cb_delete_paper(pid):
    delete_paper(pid)

def render_paper_card(p: dict, query: str = "", show_full: bool = False, selectable: bool = False):
    import html as _html
    tags = json.loads(p.get("tags") or "[]")
    status_map = {"read": ("🟢", "Read"), "reading": ("🟡", "Reading"), "unread": ("⚪", "Unread")}
    s_icon, s_label = status_map.get(p.get("status", "unread"), ("⚪", "Unread"))
    star_icon = "⭐" if p.get("starred") else "☆"

    # HTML-escape tất cả dữ liệu user
    raw_title = p.get("title") or "No Title"
    title_display = highlight_text(raw_title, query, 999) if query else _html.escape(raw_title)
    authors = _html.escape(p.get("authors") or "")
    year = _html.escape(str(p.get("year") or "—"))
    journal = _html.escape(p.get("journal") or "—")
    renamed = _html.escape(p.get("renamed_filename") or p.get("original_filename") or "—")
    pages = _html.escape(str(p.get("pages") or "—"))
    abstract = p.get("abstract") or ""

    card_class = "paper-card starred" if p.get("starred") else "paper-card"

    doi_html = f'<div style="margin-bottom:6px"><span class="doi-text">DOI: {_html.escape(p.get("doi"))}</span></div>' if p.get("doi") else ""
    tags_html = ''.join(f'<span class="tag-badge">{_html.escape(t)}</span>' for t in tags)
    project_badge = (
        f'<span style="display:inline-block;background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe;'
        f'border-radius:6px;padding:2px 10px;font-size:11px;font-weight:600;margin-bottom:6px">'
        f'🗂 {_html.escape(p.get("project",""))}</span><br>' if p.get("project") else ""
    )

    with st.container():
        if selectable:
            sel_col, card_col = st.columns([0.5, 9.5], vertical_alignment="center")
            with sel_col:
                st.checkbox("Select to delete", key=f"sel_{p['id']}", label_visibility="collapsed")
        else:
            card_col = st.container()

        with card_col:
            authors_short = authors[:80] + ('…' if len(authors) > 80 else '')
            journal_short = journal[:50] + ('…' if len(journal) > 50 else '')

            abstract_html = ""
            if abstract:
                safe_abstract = _html.escape(abstract).replace('\n', '<br>')
                preview = highlight_text(safe_abstract, query, 9999) if query else safe_abstract
                # Sử dụng trick CSS Checkbox để mở/đóng abstract mà không cần load lại pages
                toggle_id = f"toggle_abs_{p['id']}"
                abstract_html = (
                    f'<div class="abstract-container">'
                    f'<input type="checkbox" id="{toggle_id}" class="abstract-toggle">'
                    f'<div class="abstract-box">{preview}</div>'
                    f'<label for="{toggle_id}" class="abstract-label"></label>'
                    f'</div>'
                )

            card_html = (
                f'<div class="{card_class}">'
                f'<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:12px">'
                f'<div style="flex:1">'
                f'<div class="paper-title">{title_display}</div>'
                f'<div style="font-size:12px;color:#64748b;margin-bottom:6px">'
                f'👤 {authors_short} &nbsp;·&nbsp; 📅 {year} &nbsp;·&nbsp; 📖 {journal_short} &nbsp;·&nbsp; 📄 {pages} pages</div>'
                f'{project_badge}'
                f'{doi_html}'
                f'<div style="margin-bottom:8px">{tags_html}</div>'
                f'<div style="font-size:11px;font-family:\'IBM Plex Mono\',monospace;color:#94a3b8;opacity:0.8;margin-bottom:8px">🗂 {renamed}</div>'
                f'</div>'
                f'<div style="text-align:right;min-width:80px">'
                f'<div style="font-size:12px;margin-bottom:6px">{s_icon} {s_label}</div>'
                f'<div style="font-size:18px">{star_icon}</div>'
                f'</div>'
                f'</div>'
                f'{abstract_html}'
                f'</div>'
            )
            st.markdown(card_html, unsafe_allow_html=True)

            # Buttons (sử dụng callback tối ưu hiệu năng không double-rerun)
            col1, col2, col3, col4, col5, col6 = st.columns([2, 2, 2, 2, 2, 1])
            with col1:
                st.button("📖 Quick read", key=f"read_{p['id']}", use_container_width=True, on_click=cb_quick_read, args=(p["id"],))
            with col2:
                st.button("✏️ Edit", key=f"edit_{p['id']}", use_container_width=True, on_click=cb_toggle_edit, args=(p["id"],))
            with col3:
                star_lbl = "⭐ Unstar" if p.get("starred") else "☆ Star"
                st.button(star_lbl, key=f"star_{p['id']}", use_container_width=True, on_click=cb_toggle_star, args=(p["id"], p.get("starred")))
            with col4:
                current_status = p.get("status", "unread")
                status_btn_labels = {"reading": "🟡 reading", "read": "🟢 read", "unread": "⚪ Unread"}
                st.button(status_btn_labels.get(current_status, "—"), key=f"status_{p['id']}", use_container_width=True, on_click=cb_toggle_status, args=(p["id"], current_status))
            with col5:
                # ── One-click Citation ──
                if st.button("📋 Vancouver", key=f"cite_{p['id']}", use_container_width=True, help="Copy Vancouver citation"):
                    st.session_state[f"show_cite_{p['id']}"] = not st.session_state.get(f"show_cite_{p['id']}", False)
            with col6:
                st.button("🗑️", key=f"del_{p['id']}", help="Delete paper", on_click=cb_delete_paper, args=(p["id"],))

            # ── Citation popup ──
            if st.session_state.get(f"show_cite_{p['id']}", False):
                vancouver = generate_citation(p, "vancouver")
                apa       = generate_citation(p, "apa")
                chicago   = generate_citation(p, "chicago")
                with st.container():
                    st.markdown(f"""
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:16px;margin-top:6px;margin-bottom:10px">
                        <div style="font-size:11px;font-weight:700;color:#c9a84c;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:10px">📋 Copy Citation</div>
                        <div style="margin-bottom:8px">
                            <div style="font-size:10px;font-weight:600;color:#64748b;margin-bottom:4px">VANCOUVER</div>
                            <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:6px;padding:10px;font-size:12px;font-family:'IBM Plex Mono',monospace;color:#1e293b;line-height:1.6">{_html.escape(vancouver)}</div>
                        </div>
                        <div style="margin-bottom:8px">
                            <div style="font-size:10px;font-weight:600;color:#64748b;margin-bottom:4px">APA</div>
                            <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:6px;padding:10px;font-size:12px;font-family:'IBM Plex Mono',monospace;color:#1e293b;line-height:1.6">{_html.escape(apa)}</div>
                        </div>
                        <div>
                            <div style="font-size:10px;font-weight:600;color:#64748b;margin-bottom:4px">CHICAGO</div>
                            <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:6px;padding:10px;font-size:12px;font-family:'IBM Plex Mono',monospace;color:#1e293b;line-height:1.6">{_html.escape(chicago)}</div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    # Streamlit copy workaround: text_area để user có thể copy dễ dàng
                    st.text_area("📌 Vancouver (select all & copy):", value=vancouver, height=80, key=f"cite_va_{p['id']}")

    # Panel chỉnh sửa inline
    if st.session_state.edit_mode.get(p["id"]):
        render_edit_panel(p)


def render_edit_panel(p: dict):
    """Panel chỉnh sửa metadata papers."""
    with st.expander("✏️ Edit thông tin papers", expanded=True):
        c1, c2 = st.columns(2)
        with c1:
            new_title   = st.text_input("Title", value=p.get("title",""), key=f"et_{p['id']}")
            new_authors = st.text_input("Authors (separated by ;)", value=p.get("authors",""), key=f"ea_{p['id']}")
            new_year    = st.number_input("Year", min_value=1900, max_value=2100, value=int(p.get("year") or 2024), key=f"ey_{p['id']}")
        with c2:
            new_journal  = st.text_input("Journals", value=p.get("journal",""), key=f"ej_{p['id']}")
            new_doi      = st.text_input("DOI", value=p.get("doi",""), key=f"ed_{p['id']}")
            new_keywords = st.text_input("Keywords (separated by ,)", value=p.get("keywords",""), key=f"ek_{p['id']}")

        new_abstract = st.text_area("Abstract (Abstract)", value=p.get("abstract",""), height=120, key=f"eab_{p['id']}")
        new_notes    = st.text_area("Personal Notes", value=p.get("notes",""), height=80, key=f"en_{p['id']}")

        tags_str = ", ".join(json.loads(p.get("tags") or "[]"))
        new_tags_str = st.text_input("Tags (separated by ,)", value=tags_str, key=f"etg_{p['id']}")

        # ─── Project ──────────────────────────────────────────────────────────
        existing_projects = get_all_projects()
        project_choices = ["— No project —"] + existing_projects
        current_project = p.get("project") or ""
        proj_index = project_choices.index(current_project) if current_project in project_choices else 0
        col_proj1, col_proj2 = st.columns([2, 1])
        with col_proj1:
            selected_project = st.selectbox("🗂 Project", project_choices, index=proj_index, key=f"eproj_{p['id']}")
        with col_proj2:
            new_project_name = st.text_input("Or create new project", placeholder="e.g. Vietnamese Facebase", key=f"enewproj_{p['id']}")
        final_project = new_project_name.strip() if new_project_name.strip() else (selected_project if selected_project != "— No project —" else "")

        col_save, col_cancel = st.columns([1, 4])
        with col_save:
            if st.button("💾 Save", key=f"save_{p['id']}", type="primary"):
                new_tags = [t.strip() for t in new_tags_str.split(",") if t.strip()]
                updated = {
                    "title": new_title,
                    "authors": new_authors,
                    "year": int(new_year),
                    "journal": new_journal,
                    "doi": new_doi,
                    "keywords": new_keywords,
                    "abstract": new_abstract,
                    "notes": new_notes,
                    "tags": json.dumps(new_tags),
                    "project": final_project,
                }
                # Cập nhật tên file đổi dựa trên metadata mới
                updated["renamed_filename"] = build_renamed_filename({**p, **updated})
                update_paper(p["id"], updated)
                st.session_state.edit_mode[p["id"]] = False
                st.success("✅ Saved!")
                st.rerun()
        with col_cancel:
            if st.button("Cancel", key=f"cancel_{p['id']}"):
                st.session_state.edit_mode[p["id"]] = False
                st.rerun()


@st.dialog("📄 Read Fullscreen")
def show_pdf_fullscreen(pdf_b64: str, filename: str):
    st.markdown(
        """
        <style>
        /* Ép Modal Dialog mở hết cỡ toàn màn hình (100vw, 100vh) */
        div[role="dialog"] {
            width: 100vw !important;
            max-width: 100vw !important;
            height: 100vh !important;
            max-height: 100vh !important;
            margin: 0 !important;
            padding: 0 !important;
            border-radius: 0 !important;
            transform: none !important;
            top: 0 !important;
            left: 0 !important;
        }
        /* Loại bỏ padding mặc định của ruột bên trong */
        div[data-testid="stModalContent"] {
            padding: 0 !important;
        }
        div[data-testid="stModalContent"] > div {
            padding: 0 !important;
        }
        </style>
        """, unsafe_allow_html=True
    )
    pdf_iframe = f"""
        <iframe
            src="data:application/pdf;base64,{pdf_b64}#toolbar=1&navpanes=0&scrollbar=1"
            width="100%"
            height="100%"
            style="min-height: 100vh; border:none; margin: 0; padding: 0;"
            type="application/pdf">
            <p>Browser does not support PDF display.</p>
        </iframe>
    """
    st.markdown(pdf_iframe, unsafe_allow_html=True)

def render_quick_reader(paper_id: int):
    """Cửa sổ đọc nhanh papers với đầy đủ thông tin (bao gồm full_text)."""
    p_dict = get_paper_full(paper_id)
    if not p_dict:
        return
    p = p_dict
    tags = json.loads(p.get("tags") or "[]")
    highlights = json.loads(p.get("highlights") or "[]")

    # ── Header card ──
    import html as _html
    _title = _html.escape(p.get('title', '') or 'No title')
    _authors = _html.escape(p.get('authors', '') or '—')
    _year = p.get('year', '') or '—'
    _journal = _html.escape(p.get('journal', '') or '—')
    _pages = p.get('pages', '') or '—'
    _doi = _html.escape(p.get('doi', '') or '')

    doi_html = f'<span>🔗 DOI: <code style="color:#2563eb;font-size:11px">{_doi}</code></span>' if _doi else ''
    tags_html = ''.join(f'<span class="tag highlight">{_html.escape(t)}</span>' for t in tags)

    header_html = (
        '<div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:14px;padding:28px;margin-bottom:20px;box-shadow:0 4px 20px rgba(0,0,0,0.05)">'
        f'<div style="font-family:\'Lora\',serif;font-size:22px;font-weight:700;color:#1e293b;line-height:1.4;margin-bottom:12px">{_title}</div>'
        f'<div style="display:flex;flex-wrap:wrap;gap:16px;font-size:13px;color:#64748b;margin-bottom:16px">'
        f'<span>👤 {_authors}</span><span>📅 {_year}</span><span>📖 {_journal}</span><span>📄 {_pages} pages</span>{doi_html}</div>'
        f'<div>{tags_html}</div>'
        '</div>'
    )
    st.markdown(header_html, unsafe_allow_html=True)

    # Bố cục 2 cột: nội dung | ghi chú
    col_left, col_right = st.columns([3, 2])

    with col_left:
        st.markdown("#### 📝 Abstract")
        abstract = p.get("abstract") or "*(No abstract)*"
        st.markdown(f'<div class="abstract-box" style="font-size:14px;line-height:1.8">{abstract}</div>', unsafe_allow_html=True)

        if p.get("keywords"):
            st.markdown("#### 🔑 Keywords")
            kws = [k.strip() for k in p.get("keywords","").split(",") if k.strip()]
            st.markdown("&nbsp;".join(f'<span class="tag">{k}</span>' for k in kws), unsafe_allow_html=True)

        # ── PDF Viewer nhúng base64 ──
        file_path = p.get("file_path", "")
        if file_path and os.path.exists(file_path):
            c1, c2 = st.columns([12, 1])
            with c1:
                st.markdown("#### 📄 View PDF paper")
            with c2:
                import base64
                with open(file_path, "rb") as pdf_file:
                    pdf_b64 = base64.b64encode(pdf_file.read()).decode("utf-8")
                
                if st.button("⤢", key=f"fs_{paper_id}", help="View PDF Fullscreen", use_container_width=True):
                    show_pdf_fullscreen(pdf_b64, _html.escape(p.get('renamed_filename', 'paper.pdf')))
            pdf_iframe = f"""
                <iframe
                    src="data:application/pdf;base64,{pdf_b64}"
                    width="100%"
                    height="750"
                    style="border:1px solid #e2e8f0;border-radius:10px;box-shadow:0 2px 12px rgba(0,0,0,0.06)"
                    type="application/pdf">
                    <p>Browser does not support PDF display.
                    <a href="data:application/pdf;base64,{pdf_b64}" download="{_html.escape(p.get('renamed_filename','paper.pdf'))}">Download PDF</a></p>
                </iframe>
            """
            st.markdown(pdf_iframe, unsafe_allow_html=True)
        elif p.get("full_text"):
            st.markdown("#### 📄 Full-text content (Preview)")
            st.text(p["full_text"][:5000] + ("…" if len(p["full_text"]) > 5000 else ""))

    with col_right:
        st.markdown("#### 📌 Personal Notes")
        current_notes = p.get("notes") or ""
        new_notes = st.text_area(
            "Notes",
            value=current_notes,
            height=160,
            key=f"reader_notes_{paper_id}",
            label_visibility="collapsed",
            placeholder="Write observations, highlights, clinical applications...",
        )
        if st.button("💾 Save Notes", key=f"save_notes_{paper_id}"):
            update_paper(paper_id, {"notes": new_notes})
            st.success("Saved!")

        st.markdown("#### ✦ Highlights")
        new_hl = st.text_input("Add new highlight", key=f"hl_input_{paper_id}", placeholder="Paste important text here...")
        if st.button("➕ Add", key=f"hl_add_{paper_id}") and new_hl:
            highlights.append({"text": new_hl, "added": datetime.now().strftime("%Y-%m-%d")})
            update_paper(paper_id, {"highlights": json.dumps(highlights)})
            st.rerun()

        for i, hl in enumerate(highlights):
            hl_text = hl["text"] if isinstance(hl, dict) else hl
            hl_date = hl.get("added","") if isinstance(hl, dict) else ""
            st.markdown(f"""
                <div style="background:#fef9c3;border-left:3px solid #c9a84c;padding:10px 14px;
                            border-radius:0 8px 8px 0;margin-bottom:8px;font-size:13px;color:#475569;line-height:1.6">
                    "{hl_text}"
                    <div style="font-size:10px;color:#94a3b8;margin-top:4px">{hl_date}</div>
                </div>
            """, unsafe_allow_html=True)
            if st.button("✕", key=f"delhl_{paper_id}_{i}", help="Delete highlight"):
                highlights.pop(i)
                update_paper(paper_id, {"highlights": json.dumps(highlights)})
                st.rerun()

        st.markdown("#### 📊 Reading Status")
        status_opts = {"unread": 0, "reading": 1, "read": 2}
        status_disp = ["⚪ Unread", "🟡 reading", "🟢 read"]
        cur_idx = status_opts.get(p.get("status","unread"), 0)
        new_status_label = st.radio("", status_disp, index=cur_idx, key=f"status_radio_{paper_id}")
        new_status = ["unread","reading","read"][status_disp.index(new_status_label)]
        if new_status != p.get("status"):
            update_paper(paper_id, {"status": new_status})
            st.rerun()

    if st.button("← Back to library", key="back_btn"):
        st.session_state.selected_paper_id = None
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
#  APPLY FILTERS
# ══════════════════════════════════════════════════════════════════════════════
def apply_filters(papers: list[dict]) -> list[dict]:
    status_map = {"All": None, "Unread": "unread", "Reading": "reading", "Read": "read"}
    fs = status_map.get(st.session_state.get("filter_status","All"))
    fstar = st.session_state.get("filter_starred", False)
    fproject = st.session_state.get("filter_project", "All")
    result = papers
    if fs:
        result = [p for p in result if p["status"] == fs]
    if fstar:
        result = [p for p in result if p["starred"]]
    if fproject and fproject != "All":
        result = [p for p in result if (p.get("project") or "") == fproject]
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  TAB: THƯ INSTITUTE
# ══════════════════════════════════════════════════════════════════════════════
def tab_library():
    col_hdr1, col_hdr2 = st.columns([5, 1])
    with col_hdr1:
        st.markdown("## 📚 Library literature")
    with col_hdr2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("❓ Tutorial", use_container_width=True, help="Replay the onboarding tutorial"):
            conf = read_config()
            conf["has_seen_tutorial"] = False
            save_config(conf)
            st.session_state.quest_step = 1
            st.rerun()

    # Nếu đang xem chi tiết một bài
    if st.session_state.selected_paper_id:
        render_quick_reader(st.session_state.selected_paper_id)
        return

    sort_order = sort_options.get(st.session_state.get("sort_by","Recently added"), "added_at DESC")
    total_count = get_papers_count()

    # Pagination: tự bật khi thư viện > 30 papers
    PAGE_SIZE = 30
    if total_count > PAGE_SIZE:
        if "lib_page" not in st.session_state:
            st.session_state.lib_page = 0
        # Reset page khi đổi sort hoặc filter
        current_sort_key = f"{sort_order}_{st.session_state.get('filter_status','')}_{st.session_state.get('filter_starred','')}_{st.session_state.get('filter_project','')}"
        if st.session_state.get("_last_sort_key") != current_sort_key:
            st.session_state.lib_page = 0
            st.session_state["_last_sort_key"] = current_sort_key

        page_papers = get_papers_page(order=sort_order, page=st.session_state.lib_page, page_size=PAGE_SIZE)
        all_papers = page_papers
    else:
        st.session_state.lib_page = 0
        all_papers = get_all_papers(sort_order)

    papers = apply_filters(all_papers)

    if not papers:
        st.markdown("""
            <div style="text-align:center;padding:40px 20px;color:#64748b">
                <div style="font-size:56px;margin-bottom:16px">📚</div>
                <div style="font-family:'Lora',serif;font-size:22px;font-weight:700;color:#334155;margin-bottom:8px">
                    Welcome to SciKMS
                </div>
                <div style="font-size:15px;color:#64748b;margin-bottom:6px">
                    Smart Medical Literature Management System
                </div>
                <div style="font-size:13px;color:#94a3b8;max-width:500px;margin:0 auto 24px">
                    Start by uploading PDF papers below.<br>
                    The system will automatically extract metadata, DOI, tags for you.
                </div>
            </div>
        """, unsafe_allow_html=True)

        # Tích hợp Import trực tiếp ando pages chính
        st.markdown("---")
        st.markdown("#### ⬆️ Get started — Upload papers")
        uploaded_files = st.file_uploader(
            "Drag & drop PDF here or click to select (multiple files)",
            type=["pdf"],
            accept_multiple_files=True,
            key="quick_import_uploader",
        )

        if uploaded_files:
            if st.button(f"🚀 Process {len(uploaded_files)} PDF files", type="primary", use_container_width=True):
                results = []
                progress = st.progress(0)
                status_text = st.empty()
                for i, f in enumerate(uploaded_files):
                    status_text.markdown(f"**Processing:** `{f.name}` ({i+1}/{len(uploaded_files)})")
                    result = process_uploaded_pdf(f)
                    results.append({"file": f.name, "result": result})
                    progress.progress((i+1)/len(uploaded_files))
                status_text.empty()
                progress.empty()
                ok   = [r for r in results if r["result"] and "error" not in r["result"]]
                errs = [r for r in results if not r["result"] or "error" in (r["result"] or {})]
                if ok:
                    st.success(f"✅ Added **{len(ok)}** papers to the library!")
                if errs:
                    for e in errs:
                        msg = (e["result"] or {}).get("error", "Unknown error")
                        st.warning(f"⚠️ `{e['file']}`: {msg}")
                st.rerun()

        # Nút chuyển sang tab Import đầy đủ
        st.markdown("")
        col_a, col_b, col_c = st.columns([1, 2, 1])
        with col_b:
            if st.button("📂 Or open full PDF Import page →", use_container_width=True):
                st.session_state.active_tab = "⬆️ Import"
                st.rerun()
        return

    c_info, c_bulk = st.columns([1, 1])
    with c_info:
        st.markdown(f'<div style="color:#64748b;font-size:13px;margin-top:10px;margin-bottom:20px">Showing <b style="color:#c9a84c">{len(papers)}</b> / {len(all_papers)} papers</div>', unsafe_allow_html=True)
    with c_bulk:
        st.markdown('<div style="text-align:right">', unsafe_allow_html=True)
        bulk_mode = st.toggle("🗑️ Bulk selection mode", key="bulk_delete_mode")
        st.markdown('</div>', unsafe_allow_html=True)

    if bulk_mode:
        selected_ids = [p["id"] for p in papers if st.session_state.get(f"sel_{p['id']}")]
        col_del, col_delall = st.columns([1, 1])
        with col_del:
            if selected_ids:
                st.error(f"Selected {len(selected_ids)} papers. This action cannot be undone.")
                if st.button(f"🚨 Delete {len(selected_ids)} selected papers", type="primary", use_container_width=True, key="bulk_del_btn"):
                    for pid in selected_ids:
                        delete_paper(pid)
                        if f"sel_{pid}" in st.session_state:
                            del st.session_state[f"sel_{pid}"]
                    st.success(f"Deleted {len(selected_ids)} papers!")
                    st.rerun()
        with col_delall:
            if not st.session_state.get("confirm_delall_lib"):
                if st.button("🗑️ Delete ENTIRE library", use_container_width=True, key="delall_lib_btn"):
                    st.session_state["confirm_delall_lib"] = True
                    st.rerun()
            else:
                st.warning(f"⚠️ Will delete all **{len(papers)}** papers and PDF files. Cannot be undone!")
                cc1, cc2 = st.columns(2)
                with cc1:
                    if st.button("✅ Confirm delete all", type="primary", use_container_width=True, key="delall_lib_confirm"):
                        delete_all_papers()
                        st.session_state["confirm_delall_lib"] = False
                        st.success("Entire library deleted!")
                        st.rerun()
                with cc2:
                    if st.button("❌ Cancel", use_container_width=True, key="delall_lib_cancel"):
                        st.session_state["confirm_delall_lib"] = False
                        st.rerun()

    for p in papers:
        render_paper_card(p, selectable=bulk_mode)

    # ─── Pagination controls ──────────────────────────────────────────────────
    if total_count > PAGE_SIZE:
        total_pages = (total_count + PAGE_SIZE - 1) // PAGE_SIZE
        cur_page = st.session_state.lib_page
        st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
        pg_cols = st.columns([1, 2, 1])
        with pg_cols[0]:
            if cur_page > 0:
                if st.button("← Previous Page", use_container_width=True):
                    st.session_state.lib_page -= 1
                    st.rerun()
        with pg_cols[1]:
            st.markdown(
                f'<div style="text-align:center;font-size:13px;color:#64748b;padding:8px 0">' +
                f'Page <b style="color:#c9a84c">{cur_page+1}</b> / {total_pages} ' +
                f'· Total <b>{total_count}</b> papers</div>',
                unsafe_allow_html=True
            )
        with pg_cols[2]:
            if cur_page < total_pages - 1:
                if st.button("Next Page →", use_container_width=True):
                    st.session_state.lib_page += 1
                    st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
#  TAB: TÌM KIẾM KEYWORD
# ══════════════════════════════════════════════════════════════════════════════
def tab_search():
    st.markdown("## 🔍 Search Keywords")
    st.markdown("""
        <div class="info-tip" style="margin-bottom:20px">
        Full-Text Search (Full-Text Search) trong <b>title, authors, abstract, keywords, and PDF content</b>.
        Results are highlighted directly where keywords appear.
        </div>
    """, unsafe_allow_html=True)

    query = st.text_input(
        "🔎 Enter search keywords",
        value=st.session_state.search_query,
        placeholder="For example: rhinoplasty, 3D morphometrics, Farkas, local flap...",
        key="search_input_main",
    )
    st.session_state.search_query = query

    # Popular tag suggestions
    all_papers = get_all_papers()
    all_tags = []
    for p in all_papers:
        all_tags.extend(json.loads(p.get("tags") or "[]"))
    from collections import Counter
    top_tags = [t for t, _ in Counter(all_tags).most_common(12)]
    if top_tags:
        st.markdown("**Popular topics:**")
        tag_cols = st.columns(min(len(top_tags), 6))
        for i, tag in enumerate(top_tags[:6]):
            with tag_cols[i]:
                if st.button(tag, key=f"qtag_{tag}", use_container_width=True):
                    st.session_state.search_query = tag
                    st.rerun()

    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)

    if not query:
        st.markdown('<div style="color:#64748b;text-align:center;padding:40px">Enter keywords to start searching</div>', unsafe_allow_html=True)
        return

    with st.spinner("Searching..."):
        results = fts_search(query)
        results = apply_filters(results)

    if not results:
        st.warning(f"No papers found containing **{query}**")
        return

    st.success(f"✅ Found **{len(results)}** papers containing **\"{query}\"**")

    for p in results:
        render_paper_card(p, query=query)


# ══════════════════════════════════════════════════════════════════════════════
#  PASTE REFERENCE PARSER
# ══════════════════════════════════════════════════════════════════════════════
def parse_reference_list(text: str) -> list[dict]:
    """
    Parse a pasted reference list into structured dicts.
    Handles numbered formats like:
      1. Smith J, Doe A. Title of paper. Journal Name 2020;15(2):100-110. [doi: 10.xxxx/xxx]
    Returns list of dicts: {num, raw, doi, authors, title, journal, year, volume, issue, pages}
    """
    # Split by reference numbers at start of line (1., 2., ... or 1) 2) )
    ref_pattern = re.compile(r'(?:^|\n)\s*(\d{1,3})[.)]\s+', re.MULTILINE)
    parts = ref_pattern.split(text.strip())

    refs = []
    # parts = [pre_text, num1, ref1_text, num2, ref2_text, ...]
    i = 1
    while i < len(parts) - 1:
        num = parts[i].strip()
        raw = parts[i + 1].strip() if i + 1 < len(parts) else ""
        i += 2

        if not raw:
            continue

        parsed = {"num": int(num) if num.isdigit() else len(refs)+1, "raw": raw}

        # ── Extract DOI ────────────────────────────────────────────────────
        doi_match = re.search(
            r'\[doi:\s*(10\.\S+?)\]|doi[:\s]+(10\.\S+)|https?://doi\.org/(10\.\S+)',
            raw, re.IGNORECASE
        )
        if doi_match:
            parsed["doi"] = (doi_match.group(1) or doi_match.group(2) or doi_match.group(3) or "").rstrip(".,;])")
        else:
            # Fallback: bare DOI anywhere in text
            bare = re.search(r'\b(10\.\d{4,}/[^\s"\'<>\[\]]{3,})\b', raw)
            parsed["doi"] = bare.group(1).rstrip(".,;") if bare else ""

        # ── Extract year ───────────────────────────────────────────────────
        year_m = re.search(r'\b(19|20)\d{2}\b', raw)
        parsed["year"] = int(year_m.group(0)) if year_m else None

        # ── Split authors / title / journal ───────────────────────────────
        # Strategy: split on first ". " that looks like end-of-authors,
        # then next ". " is end-of-title, rest is journal info
        # Authors typically end before a capitalised sentence (the title)
        sentences = re.split(r'\.\s+(?=[A-Z])', raw, maxsplit=3)

        if len(sentences) >= 3:
            parsed["authors"] = _clean_authors(sentences[0])
            parsed["title"]   = sentences[1].strip().rstrip(".")
            parsed["journal_raw"] = sentences[2]
        elif len(sentences) == 2:
            parsed["authors"] = _clean_authors(sentences[0])
            parsed["title"]   = sentences[1].strip().rstrip(".")
            parsed["journal_raw"] = ""
        else:
            parsed["authors"] = ""
            parsed["title"]   = raw[:120]
            parsed["journal_raw"] = ""

        # ── Parse journal name + volume/issue/pages from journal_raw ──────
        j_raw = parsed.get("journal_raw", "")
        # Remove bracketed notes like [doi:...] [Medline:...]
        j_clean = re.sub(r'\[.*?\]', '', j_raw).strip()
        # Pattern: Journal Name YYYY MonthAbbr;Vol(Issue):Pages
        j_m = re.match(
            r'^(.*?)\s*(?:19|20)\d{2}[^;]*;\s*(\d+)\s*\(([^)]+)\)\s*:\s*([\d\-–]+)',
            j_clean
        )
        if j_m:
            parsed["journal"] = j_m.group(1).strip().rstrip(",.")
            parsed["volume"]  = j_m.group(2)
            parsed["issue"]   = j_m.group(3)
            parsed["pages"]   = j_m.group(4)
        else:
            # Simpler: everything before a year-like token is journal name
            j_name_m = re.match(r'^(.*?)\s+(?:19|20)\d{2}', j_clean)
            parsed["journal"] = j_name_m.group(1).strip().rstrip(",.") if j_name_m else j_clean[:80]
            parsed["volume"]  = ""
            parsed["issue"]   = ""
            # Try to grab pages x-xx or xxx-xxx
            pages_m = re.search(r':\s*([\d\-–xvi]+)', j_clean)
            parsed["pages"]   = pages_m.group(1) if pages_m else ""

        refs.append(parsed)

    return refs


def _clean_authors(raw_authors: str) -> str:
    """Normalise raw author string to 'Last F; Last F' semicolon-separated format."""
    # Remove leading number if any (e.g. "1. Smith J")
    raw_authors = re.sub(r'^\d+[.)]\s*', '', raw_authors.strip())
    # Split on ", " but keep "Last, F" intact by splitting on "; " or " and " or on pattern "X, " where X is a name
    # Simple approach: the raw string is already "Smith J, Doe A, ..." → convert commas between authors to semicolons
    # Authors in Vancouver look like: "Smith J, Doe AB, Jones CD"
    # Each author = word(s) + uppercase initials
    # We'll just return trimmed as-is but replace trailing comma
    return raw_authors.strip().rstrip(",.")


def import_from_parsed_ref(parsed: dict, auto_pdf: bool = False) -> dict:
    """
    Import a single parsed reference into DB.
    If DOI found → use import_by_doi_only (gets full metadata from Crossref).
    Otherwise → insert from parsed fields directly.
    """
    doi = parsed.get("doi", "").strip()

    # ── Route 1: DOI available → Crossref lookup ──────────────────────────
    if doi:
        result = import_by_doi_only(doi, auto_download_pdf=auto_pdf)
        # If Crossref failed, fall through to manual insert with parsed data
        if "error" not in result:
            return result
        # Crossref failed – attempt manual with what we have
        crossref_err = result.get("error", "")
    else:
        crossref_err = ""

    # ── Route 2: No DOI or Crossref failed → manual insert ────────────────
    title   = parsed.get("title", "").strip()
    authors = parsed.get("authors", "").strip()
    year    = parsed.get("year") or datetime.now().year
    journal = parsed.get("journal", "").strip()

    if not title:
        return {"error": f"Ref #{parsed.get('num','?')}: Cannot extract title. " + crossref_err}

    # Check duplicate by title
    dup = check_smart_duplicate(title, doi)
    if dup:
        return {"error": f"Đã tồn tại: «{dup['title'][:60]}»"}

    # Tìm PDF from nhiều nguồn nếu có DOI hoặc tiêu đề
    file_path = ""
    pages = 0
    full_text = f"{title} {authors} {journal}"
    pdf_source = ""
    if auto_pdf:
        oa = find_open_access_pdf(doi=doi, title=title)
        if oa["found"]:
            dl = download_and_save_pdf(oa["url"], doi=doi,
                                       filename_hint=re.sub(r"[^\w]", "_", title)[:30])
            if dl["success"]:
                file_path = dl["file_path"]
                pages     = dl.get("pages", 0)
                pdf_source = oa["source"]
                if dl.get("full_text"):
                    full_text = dl["full_text"]

    md5_src = file_path if file_path else (doi if doi else (title + str(year) + authors[:20]))
    if file_path and os.path.exists(file_path):
        with open(file_path, "rb") as _f:
            md5 = hashlib.md5(_f.read()).hexdigest()
    else:
        md5 = hashlib.md5(md5_src.encode()).hexdigest()

    paper = {
        "md5": md5,
        "original_filename": f"(ref-paste #{parsed.get('num','?')})",
        "title": title,
        "authors": authors,
        "year": int(year),
        "journal": journal,
        "doi": doi,
        "abstract": "",
        "keywords": "",
        "full_text": full_text,
        "tags": json.dumps(auto_tag("", "", f"{title} {journal}")),
        "status": "unread",
        "starred": 0,
        "pages": pages,
        "added_at": datetime.now().strftime("%Y-%m-%d"),
        "file_path": file_path,
        "notes": "",
        "highlights": "[]",
        "project": "",
        "_pdf_source": pdf_source,
    }
    paper["renamed_filename"] = build_renamed_filename(paper)

    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO papers
            (md5,original_filename,renamed_filename,title,authors,year,journal,doi,
             abstract,keywords,full_text,tags,notes,highlights,status,starred,pages,added_at,file_path,project)
            VALUES (:md5,:original_filename,:renamed_filename,:title,:authors,:year,:journal,:doi,
                    :abstract,:keywords,:full_text,:tags,:notes,:highlights,:status,:starred,:pages,:added_at,:file_path,:project)
        """, paper)
        conn.commit()
        return paper
    except sqlite3.IntegrityError:
        return {"error": "Bài báo đã tồn tại (Duplicate)."}
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════════
#  TAB: IMPORT PDF
# ══════════════════════════════════════════════════════════════════════════════
def tab_import():
    st.markdown("## ⬆️ Import")

    pipeline_html = """
    <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-bottom:28px;padding:20px;
                background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,0.03)">
        <div style="text-align:center;flex:1;min-width:90px">
            <div style="font-size:22px">📄</div>
            <div style="font-size:11px;font-family:'IBM Plex Mono';color:#c9a84c;margin-top:4px">Upload PDF</div>
        </div>
        <div style="color:#2d3748;font-size:20px">→</div>
        <div style="text-align:center;flex:1;min-width:90px">
            <div style="font-size:22px">🔎</div>
            <div style="font-size:11px;font-family:'IBM Plex Mono';color:#7eb8e0;margin-top:4px">Extract DOI</div>
        </div>
        <div style="color:#2d3748;font-size:20px">→</div>
        <div style="text-align:center;flex:1;min-width:90px">
            <div style="font-size:22px">🌐</div>
            <div style="font-size:11px;font-family:'IBM Plex Mono';color:#7eb8e0;margin-top:4px">Crossref API</div>
        </div>
        <div style="color:#2d3748;font-size:20px">→</div>
        <div style="text-align:center;flex:1;min-width:90px">
            <div style="font-size:22px">🤖</div>
            <div style="font-size:11px;font-family:'IBM Plex Mono';color:#c9a84c;font-weight:600;margin-top:4px">GEMINI AI</div>
        </div>
        <div style="color:#2d3748;font-size:20px">→</div>
        <div style="text-align:center;flex:1;min-width:90px">
            <div style="font-size:22px">🏷️</div>
            <div style="font-size:11px;font-family:'IBM Plex Mono';color:#4caf82;margin-top:4px">Auto-Tag</div>
        </div>
        <div style="color:#2d3748;font-size:20px">→</div>
        <div style="text-align:center;flex:1;min-width:90px">
            <div style="font-size:22px">✅</div>
            <div style="font-size:11px;font-family:'IBM Plex Mono';color:#4caf82;margin-top:4px">Save to DB</div>
        </div>
    </div>
    """
    st.markdown(pipeline_html, unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Drag & drop PDF here or click to select (multiple files)",
        type=["pdf"],
        accept_multiple_files=True,
        key="pdf_uploader",
    )

    if uploaded_files:
        if st.button(f"🚀 Process {len(uploaded_files)} PDF files", type="primary"):
            results = []
            progress = st.progress(0)
            status_text = st.empty()

            for i, f in enumerate(uploaded_files):
                status_text.markdown(f"**Processing:** `{f.name}` ({i+1}/{len(uploaded_files)})")
                result = process_uploaded_pdf(f)
                results.append({"file": f.name, "result": result})
                progress.progress((i+1)/len(uploaded_files))

            status_text.empty()
            progress.empty()
            st.session_state.upload_results = results

            ok   = [r for r in results if r["result"] and "error" not in r["result"]]
            errs = [r for r in results if not r["result"] or "error" in (r["result"] or {})]
            st.success(f"✅ Successfully added **{len(ok)}** papers")
            if errs:
                for e in errs:
                    msg = (e["result"] or {}).get("error", "Unknown error")
                    st.warning(f"⚠️ `{e['file']}`: {msg}")

    # ─── DOI IMPORT SECTION ─────────────────────────────────────────────────────
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    st.markdown("#### 🔬 Import papers via DOI / PMID")
    st.markdown("""
        <div class="info-tip" style="margin-bottom:16px">
        <b>No PDF file needed!</b> Just paste the DOI, the system automatically fetches metadata from
        <b>Crossref</b> and can auto-download free <b>Open Access</b> PDFs via <b>Unpaywall</b>.
        </div>
    """, unsafe_allow_html=True)

    doi_col1, doi_col2 = st.columns([3, 1])
    with doi_col1:
        doi_input = st.text_input(
            "Enter DOI",
            placeholder="e.g. 10.1016/j.bjps.2020.01.001  or  https://doi.org/10.1001/jamafacial.2021.0600",
            key="doi_import_input",
            label_visibility="collapsed",
        )
    with doi_col2:
        auto_pdf = st.checkbox("🔍 Auto-download OA PDF", value=True, key="doi_auto_pdf",
                               help="Use Unpaywall to auto-download free Open Access PDFs if available")

    # Multi-DOI bulk import
    with st.expander("📋 Import multiple DOIs (Bulk DOI)", expanded=False):
        bulk_dois_text = st.text_area(
            "Paste DOI list (one DOI per line)",
            placeholder="10.1016/j.bjps.2020.01.001\n10.1001/jamafacial.2021.0600\n...",
            height=120,
            key="bulk_doi_input",
        )
        auto_pdf_bulk = st.checkbox("🔍 Auto-download OA PDF (bulk)", value=False, key="bulk_doi_auto_pdf")
        if st.button("🚀 Import all DOIs in list", key="bulk_doi_btn"):
            raw_dois = [d.strip() for d in bulk_dois_text.splitlines() if d.strip()]
            if raw_dois:
                b_ok, b_err = 0, []
                b_progress = st.progress(0)
                b_status   = st.empty()
                for i, raw_doi in enumerate(raw_dois):
                    b_status.markdown(f"**Importing DOI {i+1}/{len(raw_dois)}:** `{raw_doi}`")
                    result = import_by_doi_only(raw_doi, auto_download_pdf=auto_pdf_bulk)
                    if "error" in result:
                        b_err.append(f"• `{raw_doi}`: {result['error']}")
                    else:
                        b_ok += 1
                    b_progress.progress((i+1)/len(raw_dois))
                b_progress.empty(); b_status.empty()
                if b_ok:
                    st.success(f"✅ Successfully imported **{b_ok}** papers from DOI!")
                for msg in b_err:
                    st.warning(msg)
                st.rerun()
            else:
                st.warning("No DOIs entered.")

    # ─── PASTE REFERENCE LIST SECTION ──────────────────────────────────────────
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    st.markdown("#### 📋 Paste & Import References")
    st.markdown("""
        <div class="info-tip" style="margin-bottom:16px">
        <b>✨ New feature!</b> Copy the entire reference list from any paper,
        paste here — the system automatically parses each reference, extracts DOI, looks up metadata
        from <b>Crossref</b> and adds it to the library. Supports formats <b>Vancouver, APA, NLM/PubMed</b>.
        </div>
    """, unsafe_allow_html=True)

    ref_text_input = st.text_area(
        "Paste reference list here",
        placeholder=(
            "1. Spear FM, Kokich VG. A multidisciplinary approach to esthetic dentistry. "
            "Dent Clin North Am 2007 Apr;51(2):487-505. [doi: 10.1016/j.cden.2006.12.007]\n"
            "2. Jazayeri HE, Kang S, et al. Advancements in craniofacial prosthesis. "
            "J Adv Prosthodont 2018 Dec;10(6):430-439. [doi: 10.4047/jap.2018.10.6.430]\n"
            "3. ..."
        ),
        height=200,
        key="ref_paste_input",
        label_visibility="collapsed",
    )

    ref_auto_pdf = st.checkbox(
        "Auto-download Open Access PDF (Unpaywall) if DOI exists",
        value=False,
        key="ref_paste_auto_pdf",
    )

    if ref_text_input and ref_text_input.strip():
        parsed_refs = parse_reference_list(ref_text_input)
        if parsed_refs:
            st.markdown(f"""
                <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;
                            padding:12px 18px;margin-bottom:12px;font-size:13px;color:#15803d">
                    Detected <b>{len(parsed_refs)}</b> references.
                    Click <b>Import</b> to add to library.
                </div>
            """, unsafe_allow_html=True)
            with st.expander("Preview parsed list", expanded=True):
                preview_rows = []
                for r in parsed_refs:
                    preview_rows.append({
                        "#": r["num"],
                        "Authors": (r.get("authors") or "")[:45],
                        "Title (detected)": (r.get("title") or "")[:60],
                        "Journal": (r.get("journal") or "")[:35],
                        "Year": r.get("year") or "—",
                        "DOI": "OK " + r["doi"][:30] if r.get("doi") else "No DOI",
                    })
                df_preview = pd.DataFrame(preview_rows)
                st.dataframe(df_preview, use_container_width=True, hide_index=True, height=280)
                doi_count = sum(1 for r in parsed_refs if r.get("doi"))
                st.caption(f"{doi_count}/{len(parsed_refs)} references have DOI (will query full Crossref)")

            if st.button(
                f"Import {len(parsed_refs)} references tham khao vao thu vien",
                type="primary",
                key="ref_paste_import_btn",
                use_container_width=True,
            ):
                ok_list, err_list = [], []
                prog = st.progress(0)
                status_ph = st.empty()
                for i, ref in enumerate(parsed_refs):
                    label = (ref.get("title") or ref.get("doi") or f"#{ref['num']}")[:55]
                    status_ph.markdown(f"**Importing [{i+1}/{len(parsed_refs)}]:** {label}")
                    result = import_from_parsed_ref(ref, auto_pdf=ref_auto_pdf)
                    if "error" in result:
                        err_list.append(f"#{ref['num']} - {result['error']}")
                    else:
                        ok_list.append(result)
                    prog.progress((i + 1) / len(parsed_refs))
                prog.empty()
                status_ph.empty()
                if ok_list:
                    pdf_saved = sum(1 for r in ok_list if r.get("file_path") or r.get("_pdf_downloaded"))
                    msg = f"Successfully added {len(ok_list)} references vao thu vien!"
                    if pdf_saved:
                        msg += f" ({pdf_saved} PDF files downloaded and saved)"
                    st.success(msg)
                    for r in ok_list[:8]:
                        has_pdf = bool(r.get("file_path") or r.get("_pdf_downloaded"))
                        pdf_badge = (
                            f'<span style="background:#dcfce7;color:#15803d;border-radius:4px;padding:2px 7px;font-size:10px;font-weight:600;margin-left:6px">' +
                            f'PDF {r.get("_pdf_source","") or "saved"}</span>'
                        ) if has_pdf else (
                            '<span style="background:#f1f5f9;color:#94a3b8;border-radius:4px;padding:2px 7px;font-size:10px">no PDF</span>'
                        )
                        st.markdown(f"""
                        <div style="display:flex;gap:8px;align-items:center;padding:8px 12px;
                                    background:#f8fafc;border-radius:8px;margin-bottom:6px;border:1px solid #e2e8f0">
                            <span style="color:#15803d;font-size:16px">✅</span>
                            <div style="flex:1">
                                <div style="font-weight:600;font-size:13px;color:#1e293b">{r.get('title','')[:80]} {pdf_badge}</div>
                                <div style="font-size:11px;color:#94a3b8">{r.get('authors','')[:50]} · {r.get('year','')} · {r.get('journal','')[:40]}</div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    if len(ok_list) > 8:
                        st.caption(f"va {len(ok_list)-8} bai khac")
                if err_list:
                    with st.expander(f"{len(err_list)} errors / duplicates", expanded=len(ok_list) == 0):
                        for msg in err_list:
                            st.warning(msg)
                if ok_list:
                    st.rerun()
        else:
            st.warning("Khong phat hien duoc references tham khao. Kiem tra dinh dang - moi references can bat dau bang so thu tu (1. 2. 3. ...)")

    # ─── PMID IMPORT SECTION ────────────────────────────────────────────────────
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    st.markdown("#### 🔬 Import papers qua PMID (PubMed ID)")
    st.markdown("""
        <div class="info-tip" style="margin-bottom:16px">
        <b>PubMed ID (PMID)</b> là số định danh trên PubMed — ví dụ: <code>17532924</code>.
        Hệ thống tự động lấy đầy đủ metadata + MeSH keywords from <b>PubMed E-utilities API</b> (miễn phí).
        </div>
    """, unsafe_allow_html=True)

    pmid_col1, pmid_col2 = st.columns([3, 1])
    with pmid_col1:
        pmid_input = st.text_input(
            "Nhập PMID",
            placeholder="Ví dụ: 17532924",
            key="pmid_import_input",
            label_visibility="collapsed",
        )
    with pmid_col2:
        pmid_auto_pdf = st.checkbox("🔍 Tải PDF OA", value=True, key="pmid_auto_pdf")

    if pmid_input and st.button("⚡ Import from PMID", type="primary", key="pmid_import_btn"):
        with st.spinner("🔍 Đang tra cứu PubMed..."):
            result = import_by_pmid(pmid_input.strip(), auto_download_pdf=pmid_auto_pdf)
        if "error" in result:
            st.error(result["error"])
        else:
            tags = json.loads(result.get("tags") or "[]")
            pdf_badge = ""
            if result.get("_pdf_downloaded"):
                pdf_badge = f'<span style="background:#dcfce7;color:#15803d;border-radius:4px;padding:2px 8px;font-size:11px;font-weight:600">PDF {result.get("_pdf_source","saved")}</span>'
            st.markdown(f"""
            <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:12px;padding:20px;margin-top:12px">
                <div style="font-size:13px;color:#15803d;font-weight:700;margin-bottom:6px">✅ Added to library! {pdf_badge}</div>
                <div style="font-weight:600;font-size:15px;color:#1e293b;margin-bottom:6px">{result.get("title","")}</div>
                <div style="font-size:12px;color:#64748b;margin-bottom:8px">
                    {result.get("authors","")} · {result.get("year","")} · {result.get("journal","")}
                </div>
                <div>{"".join(f'<span class="tag">{t}</span>' for t in tags)}</div>
            </div>
            """, unsafe_allow_html=True)
            st.rerun()

    if doi_input and st.button("Import from DOI", type="primary", key="doi_import_btn"):
        with st.spinner("🔍 Đang looks up metadata from Crossref..."):
            result = import_by_doi_only(doi_input.strip(), auto_download_pdf=auto_pdf)
        if "error" in result:
            st.error(result["error"])
        else:
            tags = json.loads(result.get("tags") or "[]")
            pdf_status = ""
            if auto_pdf:
                if result.get("_pdf_downloaded"):
                    pdf_status = '<span style="color:#15803d;font-weight:600">✅ OA PDF downloaded</span>'
                elif result.get("_pdf_url"):
                    pdf_status = f'<span style="color:#b45309">⚠️ Found PDF link but cannot download. <a href="{result["_pdf_url"]}" target="_blank">Open link →</a></span>'
                else:
                    pdf_status = '<span style="color:#94a3b8">ℹ️ No Open Access PDF found</span>'
            st.markdown(f"""
            <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:12px;padding:20px;margin-top:12px">
                <div style="font-size:13px;color:#15803d;font-weight:700;margin-bottom:6px">✅ Added to library!</div>
                <div style="font-weight:600;font-size:15px;color:#1e293b;margin-bottom:6px">{result.get('title','')}</div>
                <div style="font-size:12px;color:#64748b;margin-bottom:8px">
                    {result.get('authors','')} · {result.get('year','')} · {result.get('journal','')}
                </div>
                <div>{''.join(f'<span class="tag">{t}</span>' for t in tags)}</div>
                {f'<div style="margin-top:8px;font-size:12px">{pdf_status}</div>' if auto_pdf else ''}
            </div>
            """, unsafe_allow_html=True)
            st.rerun()

    # Latest processing results
    if st.session_state.upload_results:
        st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
        st.markdown("#### 📋 Latest import results")
        for r in st.session_state.upload_results:
            res = r["result"] or {}
            if "error" not in res and res:
                tags = json.loads(res.get("tags") or "[]")
                st.markdown(f"""
                <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:20px;margin-bottom:12px;box-shadow:0 2px 8px rgba(0,0,0,0.04)">
                    <div style="display:flex;gap:8px;margin-bottom:8px">
                        <span style="color:#15803d;font-size:13px">✅</span>
                        <b style="color:#1e293b;font-size:14px">{res.get('title','')}</b>
                    </div>
                    <div style="font-size:12px;color:#94a3b8;margin-bottom:6px">
                        {res.get('authors','')} · {res.get('year','')} · {res.get('journal','')}
                    </div>
                    <div style="font-size:11px;color:#4a5568;font-family:'IBM Plex Mono';margin-bottom:8px">
                        <span class="rename-old">{res.get('original_filename','')}</span>
                        &nbsp;→&nbsp;
                        <span class="rename-new">{res.get('renamed_filename','')}</span>
                    </div>
                    <div>{''.join(f'<span class="tag">{t}</span>' for t in tags)}</div>
                </div>
                """, unsafe_allow_html=True)

    # Thêm manual
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)

    # ─── Nút xoá toàn bộ ───────────────────────────────────────────────────
    with st.expander("⚠️ Danger Zone — Delete Data", expanded=False):
        all_count = len(get_all_papers())
        st.markdown(f"""
            <div style="background:#fff1f2;border:1px solid #fecdd3;border-radius:10px;padding:16px;margin-bottom:12px">
                <b style="color:#be123c">🗑️ Delete entire library</b><br>
                <span style="font-size:13px;color:#9f1239">Will delete <b>{all_count}</b> papers and all imported PDF files. This action cannot be undone.</span>
            </div>
        """, unsafe_allow_html=True)
        confirm = st.checkbox("I understand and want to delete all", key="confirm_delete_all")
        if confirm:
            if st.button("🚨 DELETE ALL IMMEDIATELY", type="primary", use_container_width=True, key="do_delete_all"):
                delete_all_papers()
                st.session_state.upload_results = []
                st.success(f"✅ Successfully deleted all {all_count} papers!")
                st.rerun()

    st.markdown("#### ✍️ Add manual paper (no PDF)")
    with st.expander("Add details directly"):
        c1, c2 = st.columns(2)
        with c1:
            m_title   = st.text_input("Title *", key="m_title")
            m_authors = st.text_input("Authors (separated by ;)", key="m_authors")
            m_year    = st.number_input("Year", 1900, 2100, datetime.now().year, key="m_year")
        with c2:
            m_journal  = st.text_input("Journals", key="m_journal")
            m_doi      = st.text_input("DOI", key="m_doi")
            m_keywords = st.text_input("Keywords", key="m_keywords")
        m_abstract = st.text_area("Abstract", height=100, key="m_abstract")

        if st.button("➕ Add to library", type="primary") and m_title:
            paper = {
                "md5": hashlib.md5(m_title.encode()).hexdigest(),
                "original_filename": "(manual)",
                "title": m_title,
                "authors": m_authors,
                "year": int(m_year),
                "journal": m_journal,
                "doi": m_doi,
                "abstract": m_abstract,
                "keywords": m_keywords,
                "full_text": f"{m_title} {m_abstract}",
                "tags": json.dumps(auto_tag("", m_keywords, m_abstract)),
                "status": "unread",
                "starred": 0,
                "pages": 0,
                "added_at": datetime.now().strftime("%Y-%m-%d"),
                "file_path": "",
                "notes": "",
                "highlights": "[]",
            }
            paper["renamed_filename"] = build_renamed_filename(paper)
            conn = get_db()
            try:
                conn.execute("""
                    INSERT INTO papers
                    (md5,original_filename,renamed_filename,title,authors,year,journal,doi,
                     abstract,keywords,full_text,tags,notes,highlights,status,starred,pages,added_at,file_path)
                    VALUES (:md5,:original_filename,:renamed_filename,:title,:authors,:year,:journal,:doi,
                            :abstract,:keywords,:full_text,:tags,:notes,:highlights,:status,:starred,:pages,:added_at,:file_path)
                """, paper)
                conn.commit()
                st.success("✅ Added!")
            except sqlite3.IntegrityError:
                st.warning("This paper already exists (Duplicate MD5)")
            finally:
                conn.close()


# ══════════════════════════════════════════════════════════════════════════════
#  TAB: ĐỔI TÊN FILE
# ══════════════════════════════════════════════════════════════════════════════
def tab_rename():
    st.markdown("## ✏️ Smart batch rename")
    st.markdown("""
        <div class="info-tip" style="margin-bottom:20px">
        Standard format: <code style="color:#c9a84c">[Year] Author1_Author2 - Short_Title.pdf</code><br>
        Example: <code style="color:#4caf82">[2023] Zhang_Nguyen - 3D_Photogrammetric_Rhinoplasty_Outcomes.pdf</code>
        </div>
    """, unsafe_allow_html=True)

    papers = get_all_papers()
    if not papers:
        st.info("No papers yet. Import first.")
        return

    # Preview bảng đổi tên
    st.markdown("#### 📋 Rename Preview")
    rename_data = []
    for p in papers:
        new_name = build_renamed_filename(p)
        rename_data.append({
            "ID": p["id"],
            "Original name": p.get("original_filename",""),
            "Suggested new name": new_name,
            "Title": (p.get("title") or "")[:60],
            "Year": p.get("year",""),
            "First author": (p.get("authors") or "").split(";")[0].strip()[:30],
        })

    df_rename = pd.DataFrame(rename_data)
    st.dataframe(df_rename, use_container_width=True, hide_index=True)

    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    st.markdown("#### ✏️ Customize individual files")

    for p in papers:
        col1, col2, col3 = st.columns([3, 3, 1])
        with col1:
            st.markdown(f'<div class="rename-old" style="padding:8px 0">{p.get("original_filename","")}</div>', unsafe_allow_html=True)
        with col2:
            suggested = build_renamed_filename(p)
            custom_name = st.text_input(
                "New name",
                value=p.get("renamed_filename") or suggested,
                key=f"rename_{p['id']}",
                label_visibility="collapsed",
            )
        with col3:
            if st.button("💾", key=f"saverename_{p['id']}", help="Save new name and rename file"):
                # 1. Rename files vật lý
                old_path = p.get("file_path", "")
                new_path = rename_physical_file(old_path, custom_name)
                
                # 2. Cập nhật DB
                update_paper(p["id"], {
                    "renamed_filename": custom_name,
                    "file_path": new_path # Cập nhật lại đường dẫn mới
                })
                st.success("✅")
                st.rerun()

    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        if st.button("🔄 Apply auto-rename to ALL", type="primary", use_container_width=True):
            success_count = 0
            for p in papers:
                new_name = build_renamed_filename(p)
                old_path = p.get("file_path", "")
                
                # Bỏ qua nếu tên mới đã giống tên cũ trong DB
                if p.get("renamed_filename") == new_name and old_path.endswith(new_name):
                    continue

                # 1. Rename files vật lý
                new_path = rename_physical_file(old_path, new_name)
                
                # 2. Cập nhật DB
                update_paper(p["id"], {
                    "renamed_filename": new_name,
                    "file_path": new_path
                })
                success_count += 1
                
            st.success(f"✅ Successfully renamed {success_count} files on disk and in library!")
            st.rerun()

    with col_b:
        if st.button("🔄 Rescan All Metadata", type="primary", use_container_width=True):
            success_count = 0
            for p in papers:
                first_page = p.get("full_text", "")[:3000]
                updated_meta = {}
                
                curr_title = p.get("title", "")
                curr_authors = p.get("authors", "")
                
                # Ưu tiên nhận dạng Tài liệu Vietnamese
                if re.search(r'(?i)(luận án|luận văn|khoá luận|bộ giáo dục|bộ y tế|trường đại học|phd student|chuyên khoa)', first_page):
                    vn_meta = extract_vn_thesis_meta(first_page)
                    if vn_meta.get("title") and vn_meta["title"] != curr_title:
                        updated_meta["title"] = vn_meta["title"]
                    if vn_meta.get("authors") and vn_meta["authors"] != curr_authors:
                        updated_meta["authors"] = vn_meta["authors"]
                else:
                    # Logic Bọc Hậu cho Bài Báo Quốc Tế
                    fallback_meta = heuristic_fallback_extract(first_page)
                    # Chỉ ghi đè tiêu đề nếu tiêu đề hiện tại quá ngắn/lỗi hoặc là rác
                    if fallback_meta.get("title") and (len(curr_title) < 10 or is_garbage_title(curr_title)):
                        updated_meta["title"] = fallback_meta["title"]
                    # Chỉ ghi đè author nếu đang trống hoặc là rác
                    if fallback_meta.get("authors") and (not curr_authors or is_garbage_author(curr_authors)):
                        updated_meta["authors"] = fallback_meta["authors"]

                # Nếu có tìm thấy thông tin mới cải thiện hơn
                if updated_meta:
                    merged_p = {**p, **updated_meta}
                    new_name = build_renamed_filename(merged_p)
                    old_path = merged_p.get("file_path", "")
                    
                    if old_path and os.path.exists(old_path) and not old_path.endswith(new_name):
                        new_path = rename_physical_file(old_path, new_name)
                        updated_meta["renamed_filename"] = new_name
                        updated_meta["file_path"] = new_path
                    else:
                        updated_meta["renamed_filename"] = new_name
                        
                    update_paper(p["id"], updated_meta)
                    success_count += 1

            st.success(f"✅ Scanned and updated metadata, renamed {success_count} files with wrong format!")
            st.rerun()

    with col_c:
        # Xuất bản đồ đổi tên
        csv_data = "\n".join(
            f'"{p.get("original_filename","")}" → "{build_renamed_filename(p)}"'
            for p in papers
        )
        st.download_button(
            "📥 Download file renaming mapping (.txt)",
            data=csv_data,
            file_name="rename_mapping.txt",
            mime="text/plain",
            use_container_width=True,
        )



# ══════════════════════════════════════════════════════════════════════════════
#  TAB: BẢNG TÓM TẮT
# ══════════════════════════════════════════════════════════════════════════════
def tab_summary():
    st.markdown("## 📊 Literature Summary")

    papers = apply_filters(get_all_papers())
    if not papers:
        st.info("No papers yet.")
        return

    # Stats tổng quan
    years = [p["year"] for p in papers if p.get("year")]
    all_tags = []
    for p in papers:
        all_tags.extend(json.loads(p.get("tags") or "[]"))
    from collections import Counter
    top_tags = Counter(all_tags).most_common(5)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="metric-box"><div class="metric-val">{len(papers)}</div><div class="metric-label">Total papers</div></div>', unsafe_allow_html=True)
    with c2:
        yr_range = f"{min(years)}–{max(years)}" if years else "—"
        st.markdown(f'<div class="metric-box"><div class="metric-val" style="font-size:18px">{yr_range}</div><div class="metric-label">Year range</div></div>', unsafe_allow_html=True)
    with c3:
        journals = len(set(p.get("journal","") for p in papers if p.get("journal")))
        st.markdown(f'<div class="metric-box"><div class="metric-val">{journals}</div><div class="metric-label">Journals</div></div>', unsafe_allow_html=True)
    with c4:
        with_abstract = sum(1 for p in papers if p.get("abstract"))
        st.markdown(f'<div class="metric-box"><div class="metric-val">{with_abstract}</div><div class="metric-label">With abstract</div></div>', unsafe_allow_html=True)

    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)

    # Bảng dữ liệu đầy đủ
    st.markdown("#### 📋 Detailed Data Table")

    rows = []
    for p in papers:
        tags = json.loads(p.get("tags") or "[]")
        rows.append({
            "★": "⭐" if p.get("starred") else "",
            "Renamed File": p.get("renamed_filename") or p.get("original_filename",""),
            "Title papers": p.get("title",""),
            "Authors": (p.get("authors") or "").split(";")[0].strip()[:40],
            "Year": p.get("year",""),
            "Journals": (p.get("journal") or "")[:40],
            "DOI": p.get("doi",""),
            "Keywords": (p.get("keywords") or "")[:60],
            "Tags": ", ".join(tags[:4]),
            "Pages": p.get("pages",""),
            "Status": {"read":"🟢 Read","reading":"🟡 Reading","unread":"⚪ Unread"}.get(p.get("status","unread"),"—"),
            "Notes": (p.get("notes") or "")[:50],
        })

    df = pd.DataFrame(rows)

    # Filter trong bảng
    search_table = st.text_input("🔎 Filter in table", placeholder="Search by title, authors, keywords...", key="table_filter")
    if search_table:
        mask = df.apply(lambda col: col.astype(str).str.contains(search_table, case=False)).any(axis=1)
        df = df[mask]
        st.caption(f"Showing {len(df)} / {len(rows)} rows")

    st.dataframe(df, use_container_width=True, hide_index=True, height=500)

    # Abstract panel
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    st.markdown("#### 📖 Quick read abstract")

    titles = [f"{p.get('year','')} — {(p.get('title') or '')[:70]}" for p in papers]
    selected_title = st.selectbox("Select paper to view abstract", ["— Select —"] + titles, key="abs_select")
    if selected_title != "— Select —":
        idx = titles.index(selected_title)
        p = papers[idx]
        tags = json.loads(p.get("tags") or "[]")
        st.markdown(f"""
            <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:24px;margin-top:16px;box-shadow:0 4px 15px rgba(0,0,0,0.05)">
                <div style="font-family:'Lora',serif;font-size:17px;font-weight:700;color:#1e293b;margin-bottom:8px">
                    {p.get('title','')}
                </div>
                <div style="font-size:12px;color:#94a3b8;margin-bottom:12px">
                    {p.get('authors','')} · {p.get('year','')} · {p.get('journal','')}
                </div>
                {''.join(f'<span class="tag highlight">{t}</span>' for t in tags)}
                <div class="abstract-box" style="margin-top:14px">{p.get('abstract') or '<i>No abstract</i>'}</div>
                {'<div style="font-size:12px;color:#7eb8e0;margin-top:10px">🔑 <b>Keywords:</b> ' + p.get("keywords","") + '</div>' if p.get("keywords") else ''}
                {'<div style="font-size:12px;color:#e2e8f0;margin-top:10px;padding:10px;background:#1e2a3a;border-radius:8px">📌 <b>Notes:</b> ' + p.get("notes","") + '</div>' if p.get("notes") else ''}
            </div>
        """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB: XUẤT DỮ LIỆU
# ══════════════════════════════════════════════════════════════════════════════
def tab_export():
    st.markdown("## 📤 Export")

    papers = get_all_papers()
    if not papers:
        st.info("No data to export.")
        return

    # Chọn tập con
    st.markdown("#### Select papers to export")
    export_scope = st.radio(
        "",
        ["All papers", "Read papers only", "Starred papers only ⭐", "Filter by tag"],
        horizontal=True,
        key="export_scope",
    )

    if export_scope == "Read papers only":
        target = [p for p in papers if p["status"] == "read"]
    elif export_scope == "Starred papers only ⭐":
        target = [p for p in papers if p["starred"]]
    elif export_scope == "Filter by tag":
        all_tags = list(set(t for p in papers for t in json.loads(p.get("tags") or "[]")))
        selected_tag = st.selectbox("Select tag", ["—"] + sorted(all_tags), key="export_tag")
        target = [p for p in papers if selected_tag != "—" and selected_tag in json.loads(p.get("tags") or "[]")]
    else:
        target = papers

    st.markdown(f'<div style="color:#94a3b8;font-size:13px;margin-bottom:20px">Will export <b style="color:#c9a84c">{len(target)}</b> papers</div>', unsafe_allow_html=True)

    # Các định dạng xuất
    c1, c2, c3 = st.columns(3)

    with c1:
        st.markdown("""
            <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:20px;text-align:center;margin-bottom:12px;box-shadow:0 2px 8px rgba(0,0,0,0.04)">
                <div style="font-size:32px">📗</div>
                <div style="font-weight:600;color:#1e293b;margin:8px 0">RIS Format</div>
                <div style="font-size:12px;color:#64748b">EndNote · Zotero · Mendeley</div>
            </div>
        """, unsafe_allow_html=True)
        st.download_button(
            "📥 Download .RIS",
            data=export_ris(target),
            file_name=f"scikms_export_{datetime.now().strftime('%Y%m%d')}.ris",
            mime="text/plain",
            use_container_width=True,
        )

    with c2:
        st.markdown("""
            <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:20px;text-align:center;margin-bottom:12px;box-shadow:0 2px 8px rgba(0,0,0,0.04)">
                <div style="font-size:32px">📘</div>
                <div style="font-weight:600;color:#1e293b;margin:8px 0">BibTeX Format</div>
                <div style="font-size:12px;color:#64748b">LaTeX · Overleaf · Pandoc</div>
            </div>
        """, unsafe_allow_html=True)
        st.download_button(
            "📥 Download .BIB",
            data=export_bib(target),
            file_name=f"scikms_export_{datetime.now().strftime('%Y%m%d')}.bib",
            mime="text/plain",
            use_container_width=True,
        )

    with c3:
        st.markdown("""
            <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:20px;text-align:center;margin-bottom:12px;box-shadow:0 2px 8px rgba(0,0,0,0.04)">
                <div style="font-size:32px">📊</div>
                <div style="font-weight:600;color:#1e293b;margin:8px 0">Excel Report</div>
                <div style="font-size:12px;color:#64748b">Full summary table</div>
            </div>
        """, unsafe_allow_html=True)
        excel_bytes = export_excel(target)
        ext = "xlsx" if HAS_OPENPYXL else "csv"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if HAS_OPENPYXL else "text/csv"
        st.download_button(
            f"📥 Download .{ext.upper()}",
            data=excel_bytes,
            file_name=f"scikms_yvan_{datetime.now().strftime('%Y%m%d')}.{ext}",
            mime=mime,
            use_container_width=True,
        )

    # Xuất ZIP kèm PDF gốc
    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)
    st.markdown("#### 📦 Export ZIP (Renamed PDFs + metadata)")
    if st.button("🗜️ Create ZIP file", use_container_width=False):
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            # Thêm file metadata
            zf.writestr("metadata.json", json.dumps(target, ensure_ascii=False, indent=2))
            zf.writestr("bibliography.ris", export_ris(target))
            zf.writestr("bibliography.bib", export_bib(target))
            # Thêm PDF với tên đã đổi
            for p in target:
                fp = p.get("file_path","")
                if fp and os.path.exists(fp):
                    new_name = p.get("renamed_filename") or p.get("original_filename","paper.pdf")
                    zf.write(fp, f"papers/{new_name}")
        zip_buf.seek(0)
        st.download_button(
            "📥 Download ZIP",
            data=zip_buf.getvalue(),
            file_name=f"scikms_archive_{datetime.now().strftime('%Y%m%d')}.zip",
            mime="application/zip",
        )



# ══════════════════════════════════════════════════════════════════════════════
#  TAB: SETTINGS
# ══════════════════════════════════════════════════════════════════════════════
def tab_settings():
    st.markdown("## ⚙️ Settings")

    # ─── Tag Dictionary Editor ────────────────────────────────────────────────
    st.markdown("### 🏷️ Custom Auto-Tag Dictionary")
    st.markdown("""
        <div class="info-tip" style="margin-bottom:16px">
        The system will automatically assign these tags to papers during import if the keyword appears
        in the title, abstract or keywords. One term per line.
        The default dictionary focuses on <b>Plastic & Reconstructive Surgery</b> —
        you can completely replace it for other specialties.
        </div>
    """, unsafe_allow_html=True)

    current_dict = get_custom_tag_dict()
    current_text = "\n".join(current_dict)

    col_left, col_right = st.columns([3, 2])
    with col_left:
        new_dict_text = st.text_area(
            "Tag list (one term/phrase per line)",
            value=current_text,
            height=400,
            key="tag_dict_editor",
            label_visibility="collapsed",
            placeholder="Rhinoplasty\nMeta-analysis\nDeep Learning\n...",
        )
        col_save, col_reset = st.columns(2)
        with col_save:
            if st.button("💾 Save dictionary", type="primary", use_container_width=True):
                new_terms = [t.strip() for t in new_dict_text.splitlines() if t.strip()]
                if new_terms:
                    save_custom_tag_dict(new_terms)
                    st.success(f"✅ Saved {len(new_terms)} terms to dictionary!")
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.warning("Dictionary is empty — not saved.")
        with col_reset:
            if st.button("🔄 Reset to default", use_container_width=True):
                save_custom_tag_dict(_DEFAULT_TAG_DICT)
                st.success("✅ Reset to default dictionary!")
                st.rerun()

    with col_right:
        st.markdown("**📊 Current dictionary stats**")
        st.markdown(f"""
            <div class="metric-box" style="margin-bottom:12px">
                <div class="metric-val">{len(current_dict)}</div>
                <div class="metric-label">Terms in dictionary</div>
            </div>
        """, unsafe_allow_html=True)

        # Thống kê top tags đang được dùng
        papers = get_all_papers()
        if papers:
            from collections import Counter
            all_tags = []
            for p in papers:
                all_tags.extend(json.loads(p.get("tags") or "[]"))
            top = Counter(all_tags).most_common(10)
            if top:
                st.markdown("**🔥 Top used tags in library:**")
                for tag, count in top:
                    st.markdown(f"""
                        <div style="display:flex;justify-content:space-between;align-items:center;
                                    padding:5px 10px;background:#f8fafc;border-radius:6px;margin-bottom:4px;
                                    border:1px solid #e2e8f0">
                            <span style="font-size:12px;color:#1e293b">{tag}</span>
                            <span style="font-size:11px;font-family:'IBM Plex Mono',monospace;
                                         color:#c9a84c;font-weight:600">{count}×</span>
                        </div>
                    """, unsafe_allow_html=True)

        st.markdown("**💡 Suggestions by specialty:**")
        suggestions = {
            "Internal Medicine": ["Hypertension", "Atrial Fibrillation", "Heart Failure", "COPD", "Diabetes Mellitus"],
            "Oncology": ["Chemotherapy", "Immunotherapy", "Tumor Marker", "Metastasis", "Radiotherapy"],
            "Pediatrics": ["Pediatric", "Neonatal", "Growth Chart", "Vaccination", "Developmental"],
            "Surgery": ["Laparoscopy", "Anastomosis", "Complication Rate", "Intraoperative", "Postoperative"],
        }
        selected_spec = st.selectbox("Specialty", list(suggestions.keys()), key="spec_suggest")
        if st.button("➕ Add suggestions to dictionary", key="add_suggestions"):
            existing = set(get_custom_tag_dict())
            new_terms = [t for t in suggestions[selected_spec] if t not in existing]
            if new_terms:
                updated = get_custom_tag_dict() + new_terms
                save_custom_tag_dict(updated)
                st.success(f"✅ Added {len(new_terms)} new terms!")
                st.rerun()
            else:
                st.info("All these terms are already in the dictionary.")

    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)

    # ─── Re-tag toàn bộ thư viện ─────────────────────────────────────────────
    st.markdown("### 🔄 Re-tag entire library")
    st.markdown("""
        <div class="info-tip" style="margin-bottom:16px">
        After updating the dictionary, click this button to re-apply new tags to
        <b>all</b> papers existing in the library.
        </div>
    """, unsafe_allow_html=True)
    if st.button("🏷️ Re-tag entire library with new dictionary", type="primary"):
        papers = get_all_papers()
        updated = 0
        prog = st.progress(0)
        for i, p in enumerate(papers):
            combined_text = f"{p.get('full_text','')[:3000]} {p.get('abstract','')} {p.get('keywords','')}"
            new_tags = auto_tag(combined_text, p.get("keywords",""), p.get("abstract",""))
            update_paper(p["id"], {"tags": json.dumps(new_tags)})
            updated += 1
            prog.progress((i+1)/len(papers))
        prog.empty()
        st.success(f"✅ Re-tagged {updated} papers!")
        st.cache_data.clear()
        st.rerun()

    st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)

    # ─── Cache & DB info ──────────────────────────────────────────────────────
    st.markdown("### 🗄️ Database & Cache")
    col_a, col_b = st.columns(2)
    with col_a:
        total = get_papers_count()
        db_size = os.path.getsize(DB_PATH) / 1024 if os.path.exists(DB_PATH) else 0
        storage_size = sum(
            os.path.getsize(f) for f in STORAGE_DIR.glob("*.pdf") if f.is_file()
        ) / (1024*1024)
        st.markdown(f"""
            <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:20px">
                <div style="font-size:13px;font-weight:700;color:#334155;margin-bottom:12px">📊 Info</div>
                <div style="font-size:13px;color:#475569;margin-bottom:6px">📚 Total papers: <b>{total}</b></div>
                <div style="font-size:13px;color:#475569;margin-bottom:6px">🗄️ DB size: <b>{db_size:.1f} KB</b></div>
                <div style="font-size:13px;color:#475569">📁 PDF storage: <b>{storage_size:.1f} MB</b></div>
            </div>
        """, unsafe_allow_html=True)
    with col_b:
        if st.button("🔄 Clear Streamlit cache", use_container_width=True):
            st.cache_data.clear()
            st.success("✅ Cache cleared!")
        st.markdown("""
            <div style="background:#fffbeb;border:1px solid #fde68a;border-radius:10px;
                        padding:12px;font-size:12px;color:#92400e;margin-top:8px">
            ⚠️ Cache auto-expires after 30 seconds. Clear manually if you need to see changes immediately.
            </div>
        """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTING + TOP NAVIGATION BAR
# ══════════════════════════════════════════════════════════════════════════════
sort_options = {
    "Recently added": "added_at DESC",
    "Year (newest)": "year DESC",
    "Title A→Z": "title ASC",
    "Authors A→Z": "authors ASC",
}

# ─── Thanh điều hướng ngang ở đầu pages ───────────────────────────────────────
nav_tabs = ["📚 Library", "⬆️ Import", "🔍 Search", "✏️ Rename ", "📊 Summary", "📤 Export", "⚙️ Settings"]
nav_keys = ["topnav_lib", "topnav_import", "topnav_search", "topnav_rename", "topnav_summary", "topnav_export", "topnav_settings"]
nav_cols = st.columns(len(nav_tabs))
for i, nav_tab in enumerate(nav_tabs):
    with nav_cols[i]:
        is_active = st.session_state.active_tab == nav_tab
        if st.button(
            nav_tab,
            key=nav_keys[i],
            use_container_width=True,
            type="primary" if is_active else "secondary",
        ):
            st.session_state.active_tab = nav_tab
            st.rerun()

st.markdown('<hr class="custom-divider">', unsafe_allow_html=True)

# Hiển thị thanh Quest Tracker
render_quest_tracker()

tab = st.session_state.active_tab
if tab == "📚 Library":
    tab_library()
elif tab == "🔍 Search":
    tab_search()
elif tab == "⬆️ Import":
    tab_import()
elif tab == "✏️ Rename ":
    tab_rename()
elif tab == "📊 Summary":
    tab_summary()
elif tab == "📤 Export":
    tab_export()
elif tab == "⚙️ Settings":
    tab_settings()
